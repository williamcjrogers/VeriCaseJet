"""Evidence Repository API Services"""

import uuid
import logging
import re
from email.utils import parseaddr
from datetime import datetime, date, timedelta
from typing import Any

from fastapi import HTTPException
from sqlalchemy import or_, and_, func, desc
from sqlalchemy.orm import Session

from ..models import (
    Case,
    Project,
    User,
    EmailMessage,
    EvidenceItem,
    EvidenceCollection,
    EvidenceCorrespondenceLink,
    EvidenceRelation,
    EvidenceCollectionItem,
    EvidenceType,
    DocumentCategory,
    CorrespondenceLinkType,
    EvidenceRelationType,
    EmailAttachment,
    Stakeholder,
)
from ..storage import presign_put, presign_get, put_object, s3, get_object
from ..storage import S3AccessError, S3BucketError
from ..config import settings
from ..cache import get_cached, set_cached
from .categorization import infer_document_category
from .utils import (
    _safe_dict_list,
    get_file_type,
    compute_file_hash,
    log_activity,
    _apply_ag_filters,
    _apply_ag_sorting,
    get_default_user,
    EvidenceUploadInitResponse,
    EvidenceListResponse,
    ServerSideEvidenceResponse,
    CollectionSummary,
    EvidenceItemDetail,
    EvidenceItemSummary,
    AutoCategorizeRequest,
    AutoCategorizeResponse,
)

logger = logging.getLogger(__name__)

_FREE_EMAIL_DOMAINS = {
    "gmail.com",
    "googlemail.com",
    "outlook.com",
    "hotmail.com",
    "live.com",
    "msn.com",
    "yahoo.com",
    "ymail.com",
    "aol.com",
    "icloud.com",
    "me.com",
    "proton.me",
    "protonmail.com",
}

_MULTIPART_SUFFIXES = {
    "co.uk",
    "org.uk",
    "ac.uk",
    "gov.uk",
    "com.au",
    "net.au",
    "org.au",
    "co.nz",
    "com.sg",
    "com.hk",
    "com.cn",
    "com.br",
    "com.mx",
    "co.za",
    "co.in",
    "com.in",
}


def _company_from_domain(domain: str | None) -> str | None:
    if not domain:
        return None
    domain = domain.strip().lower().strip(".")
    if not domain:
        return None
    parts = [p for p in domain.split(".") if p]
    if len(parts) < 2:
        return None
    if parts[0] in {"mail", "smtp", "mx", "email", "mailer", "outbound", "inbound"}:
        parts = parts[1:] if len(parts) > 2 else parts
    suffix = ".".join(parts[-2:]) if len(parts) >= 2 else ""
    if len(parts) >= 3 and suffix in _MULTIPART_SUFFIXES:
        base = parts[-3]
    else:
        base = parts[-2]
    base = base.replace("_", " ").replace("-", " ").strip()
    if not base:
        return None
    if base.isupper():
        return base
    if len(base) <= 3:
        return base.upper()
    return " ".join(w.capitalize() for w in base.split())


def _normalize_recipient_list(values: list[str] | str | None) -> list[str]:
    if not values:
        return []
    if isinstance(values, list):
        return [v for v in values if v]
    if isinstance(values, str):
        return [v.strip() for v in re.split(r"[;,]", values) if v.strip()]
    return []


def _extract_company_parties(
    sender_email: str | None,
    recipients_to: list[str] | str | None,
    recipients_cc: list[str] | str | None,
    stakeholder_domain_map: dict[str, str],
    stakeholder_email_map: dict[str, str],
) -> list[str]:
    addresses: list[str] = []
    if sender_email:
        addresses.append(sender_email)
    addresses.extend(_normalize_recipient_list(recipients_to))
    addresses.extend(_normalize_recipient_list(recipients_cc))
    parties: list[str] = []
    seen: set[str] = set()
    for raw in addresses:
        if not raw:
            continue
        _, addr = parseaddr(raw)
        candidate = (addr or raw).strip()
        if "@" not in candidate:
            continue
        email_addr = candidate.lower()
        company = stakeholder_email_map.get(email_addr)
        domain = email_addr.split("@")[-1] if "@" in email_addr else ""
        if not company:
            company = stakeholder_domain_map.get(domain)
        if not company:
            if domain in _FREE_EMAIL_DOMAINS and domain not in stakeholder_domain_map:
                continue
            company = _company_from_domain(domain)
        if not company:
            continue
        company_key = company.strip().lower()
        if not company_key or company_key in seen:
            continue
        seen.add(company_key)
        parties.append(company)
    return parties


def _merge_parties(existing: list[str] | None, incoming: list[str]) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for name in existing or []:
        key = name.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        merged.append(name)
    for name in incoming:
        key = name.strip().lower()
        if not key or key in seen:
            continue
        seen.add(key)
        merged.append(name)
    return merged


def _resolve_document_date(
    primary: datetime | date | None, fallback: datetime | date | None
) -> datetime | date | None:
    value = primary or fallback
    if isinstance(value, datetime):
        return value.date()
    return value


def _parse_metadata_date(value: object) -> datetime | date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        if text.endswith("Z"):
            text = text.replace("Z", "+00:00")
        try:
            return datetime.fromisoformat(text)
        except ValueError:
            pass
        for fmt in [
            "%Y-%m-%dT%H:%M:%S%z",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%Y:%m:%d %H:%M:%S",
            "%d/%m/%Y %H:%M:%S",
            "%a, %d %b %Y %H:%M:%S %z",
            "%a, %d %b %Y %H:%M:%S GMT",
        ]:
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                continue
    return None


def _extract_metadata_date(metadata: dict[str, Any]) -> datetime | date | None:
    for field in ["created_date", "modified_date", "date_taken", "email_date"]:
        if field not in metadata:
            continue
        parsed = _parse_metadata_date(metadata.get(field))
        if parsed is not None:
            return parsed
    return None


def _parse_filename_date(filename: str | None) -> datetime | date | None:
    if not filename:
        return None
    name = filename.strip()
    if not name:
        return None
    patterns = [
        r"(?P<year>20\d{2})[._-](?P<month>\d{1,2})[._-](?P<day>\d{1,2})",
        r"(?P<day>\d{1,2})[._-](?P<month>\d{1,2})[._-](?P<year>20\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, name)
        if not match:
            continue
        try:
            year = int(match.group("year"))
            month = int(match.group("month"))
            day = int(match.group("day"))
            return datetime(year, month, day)
        except Exception:
            continue
    return None


def _extract_email_date(email: EmailMessage | None) -> datetime | date | None:
    if not email:
        return None
    if email.date_sent:
        return email.date_sent
    if email.date_received:
        return email.date_received
    meta = email.meta or {}
    for key in ["date", "sent_on", "sent_at", "received_on", "received_at"]:
        parsed = _parse_metadata_date(meta.get(key))
        if parsed is not None:
            return parsed
    return None


def _extract_email_preview(email: EmailMessage | None) -> str | None:
    if not email:
        return None
    preview = email.body_preview or email.body_text or email.body_html
    if not preview:
        return None
    if email.body_html and preview == email.body_html:
        preview = re.sub(r"<[^>]+>", " ", preview)
    preview = " ".join(preview.split())
    return preview[:240] if len(preview) > 240 else preview


def _load_stakeholder_maps(db, project_id, case_id):
    domain_map: dict[str, str] = {}
    email_map: dict[str, str] = {}
    query = db.query(Stakeholder)
    if project_id and case_id:
        query = query.filter(
            or_(Stakeholder.project_id == project_id, Stakeholder.case_id == case_id)
        )
    elif project_id:
        query = query.filter(Stakeholder.project_id == project_id)
    elif case_id:
        query = query.filter(Stakeholder.case_id == case_id)
    stakeholders = query.all()
    for s in stakeholders:
        org = (s.organization or s.name or "").strip()
        if not org:
            continue
        if s.email:
            email_map[s.email.lower()] = org
        if s.email_domain:
            domain_map[s.email_domain.lower()] = org
    return domain_map, email_map


async def init_evidence_upload_service(request, db):
    evidence_id = str(uuid.uuid4())
    s3_bucket = settings.S3_BUCKET or settings.MINIO_BUCKET
    date_prefix = datetime.now().strftime("%Y/%m")
    safe_filename = request.filename.replace(" ", "_")
    s3_key = f"evidence/{date_prefix}/{evidence_id}/{safe_filename}"
    content_type = request.content_type or "application/octet-stream"
    upload_url = presign_put(s3_key, content_type, expires=14400, bucket=s3_bucket)
    logger.info(f"Initiated evidence upload: {evidence_id}")
    return EvidenceUploadInitResponse(
        evidence_id=evidence_id,
        upload_url=upload_url,
        s3_bucket=s3_bucket,
        s3_key=s3_key,
    )


async def complete_evidence_upload_service(request, db):
    from ..evidence_metadata import extract_evidence_metadata

    user = get_default_user(db)
    s3_bucket = settings.S3_BUCKET or settings.MINIO_BUCKET
    existing = (
        db.query(EvidenceItem)
        .filter(EvidenceItem.file_hash == request.file_hash)
        .first()
    )
    is_duplicate = existing is not None
    duplicate_of_id = existing.id if existing else None

    # If a user is operating in Case context (case_id set) and the Case is linked
    # to a Project, mirror the Project on the evidence item so project-scoped tools
    # (AI refine / analysis / correspondence) can see the same evidence.
    resolved_project_id = request.project_id
    if request.case_id and not resolved_project_id:
        try:
            case_uuid = uuid.UUID(str(request.case_id))
            case = db.query(Case).filter(Case.id == case_uuid).first()
            if case and getattr(case, "project_id", None):
                resolved_project_id = str(case.project_id)
        except Exception:
            resolved_project_id = None

    evidence_item = EvidenceItem(
        filename=request.filename,
        file_type=get_file_type(request.filename),
        mime_type=request.mime_type,
        file_size=request.file_size,
        file_hash=request.file_hash,
        s3_bucket=s3_bucket,
        s3_key=request.s3_key,
        evidence_type=request.evidence_type,
        title=request.title or request.filename,
        description=request.description,
        document_date=request.document_date,
        manual_tags=request.tags or [],
        processing_status="pending",
        source_type="direct_upload",
        case_id=uuid.UUID(request.case_id) if request.case_id else None,
        project_id=uuid.UUID(resolved_project_id) if resolved_project_id else None,
        collection_id=(
            uuid.UUID(request.collection_id) if request.collection_id else None
        ),
        is_duplicate=is_duplicate,
        duplicate_of_id=duplicate_of_id,
        uploaded_by=user.id,
    )
    db.add(evidence_item)
    db.commit()
    db.refresh(evidence_item)
    log_activity(
        db,
        "upload",
        user.id,
        evidence_item_id=evidence_item.id,
        details={"filename": request.filename, "size": request.file_size},
    )
    db.commit()
    try:
        metadata = await extract_evidence_metadata(
            evidence_item.s3_key, evidence_item.s3_bucket
        )
        evidence_item.extracted_metadata = metadata
        evidence_item.metadata_extracted_at = datetime.now()
        mime_raw = metadata.get("mime_type")
        if mime_raw is not None:
            mime = str(mime_raw)
            evidence_item.mime_type = mime
            if mime.startswith("image/"):
                evidence_item.evidence_type = "image"
            elif mime == "application/pdf":
                evidence_item.evidence_type = "pdf"
            elif "word" in mime or mime.endswith(".document"):
                evidence_item.evidence_type = "word_document"
            elif "excel" in mime or "spreadsheet" in mime:
                evidence_item.evidence_type = "spreadsheet"
        if metadata.get("title") and not evidence_item.title:
            evidence_item.title = metadata["title"]
        if metadata.get("author") and not evidence_item.author:
            evidence_item.author = metadata["author"]
        if metadata.get("page_count") and not evidence_item.page_count:
            evidence_item.page_count = metadata["page_count"]
        doc_date_val = _extract_metadata_date(metadata)
        if doc_date_val:
            if isinstance(doc_date_val, datetime):
                evidence_item.document_date = (
                    doc_date_val.replace(tzinfo=None)
                    if doc_date_val.tzinfo
                    else doc_date_val
                )
            else:
                evidence_item.document_date = datetime.combine(
                    doc_date_val, datetime.min.time()
                )
        evidence_item.processing_error = None
        evidence_item.processing_status = "processed"
        db.commit()
    except Exception as e:
        evidence_item.processing_status = "error"
        evidence_item.processing_error = str(e)
        evidence_item.extracted_metadata = {
            "extraction_status": "error",
            "extraction_error": str(e),
            "filename": evidence_item.filename,
            "mime_type": evidence_item.mime_type,
        }
        db.commit()
        logger.error(f"Auto-extraction failed for {evidence_item.id}: {e}")
    logger.info(f"Created evidence item: {evidence_item.id}")
    return {
        "id": str(evidence_item.id),
        "filename": evidence_item.filename,
        "is_duplicate": is_duplicate,
        "duplicate_of_id": str(duplicate_of_id) if duplicate_of_id else None,
        "processing_status": evidence_item.processing_status,
        "message": "Evidence uploaded successfully",
    }


async def direct_upload_evidence_service(
    file, db, case_id, project_id, collection_id, evidence_type, tags
):
    from ..evidence_metadata import extract_evidence_metadata

    user = get_default_user(db)
    if not file.filename:
        raise HTTPException(400, "Missing filename")
    content = await file.read()
    file_hash = compute_file_hash(content)
    file_size = len(content)
    existing = (
        db.query(EvidenceItem).filter(EvidenceItem.file_hash == file_hash).first()
    )
    is_duplicate = existing is not None
    duplicate_of_id = existing.id if existing else None
    s3_bucket = settings.S3_BUCKET or settings.MINIO_BUCKET
    evidence_id = str(uuid.uuid4())
    date_prefix = datetime.now().strftime("%Y/%m")
    safe_filename = file.filename.replace(" ", "_")
    s3_key = f"evidence/{date_prefix}/{evidence_id}/{safe_filename}"
    try:
        put_object(
            key=s3_key,
            data=content,
            content_type=file.content_type or "application/octet-stream",
            bucket=s3_bucket,
        )
    except (S3BucketError, S3AccessError) as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except Exception as exc:
        logger.error("Failed to upload evidence to storage: %s", exc)
        raise HTTPException(
            status_code=500, detail="Failed to upload document to storage"
        ) from exc
    tag_list = []
    if tags:
        tag_list = [t.strip() for t in tags.split(",") if t.strip()]

    # Mirror linked project_id when operating in case context (see note in complete service).
    resolved_project_id = project_id
    if case_id and not resolved_project_id:
        try:
            case_uuid = uuid.UUID(str(case_id))
            case = db.query(Case).filter(Case.id == case_uuid).first()
            if case and getattr(case, "project_id", None):
                resolved_project_id = str(case.project_id)
        except Exception:
            resolved_project_id = None
    evidence_item = EvidenceItem(
        id=uuid.UUID(evidence_id),
        filename=file.filename,
        file_type=get_file_type(file.filename),
        mime_type=file.content_type,
        file_size=file_size,
        file_hash=file_hash,
        s3_bucket=s3_bucket,
        s3_key=s3_key,
        evidence_type=evidence_type,
        title=file.filename,
        manual_tags=tag_list,
        processing_status="pending",
        source_type="direct_upload",
        case_id=uuid.UUID(case_id) if case_id else None,
        project_id=uuid.UUID(resolved_project_id) if resolved_project_id else None,
        collection_id=uuid.UUID(collection_id) if collection_id else None,
        is_duplicate=is_duplicate,
        duplicate_of_id=duplicate_of_id,
        uploaded_by=user.id,
    )
    db.add(evidence_item)
    log_activity(
        db,
        "upload",
        user.id,
        evidence_item_id=evidence_item.id,
        details={"filename": file.filename, "size": file_size},
    )
    db.commit()
    db.refresh(evidence_item)
    try:
        metadata = await extract_evidence_metadata(
            evidence_item.s3_key, evidence_item.s3_bucket
        )
        evidence_item.extracted_metadata = metadata
        evidence_item.metadata_extracted_at = datetime.now()
        mime_raw = metadata.get("mime_type")
        if mime_raw is not None:
            mime = str(mime_raw)
            evidence_item.mime_type = mime
            if mime.startswith("image/"):
                evidence_item.evidence_type = "image"
            elif mime == "application/pdf":
                evidence_item.evidence_type = "pdf"
            elif "word" in mime or mime.endswith(".document"):
                evidence_item.evidence_type = "word_document"
            elif "excel" in mime or "spreadsheet" in mime:
                evidence_item.evidence_type = "spreadsheet"
        if metadata.get("title") and not evidence_item.title:
            evidence_item.title = metadata["title"]
        if metadata.get("author") and not evidence_item.author:
            evidence_item.author = metadata["author"]
        if metadata.get("page_count") and not evidence_item.page_count:
            evidence_item.page_count = metadata["page_count"]
        doc_date_val = _extract_metadata_date(metadata)
        if doc_date_val:
            if isinstance(doc_date_val, datetime):
                evidence_item.document_date = (
                    doc_date_val.replace(tzinfo=None)
                    if doc_date_val.tzinfo
                    else doc_date_val
                )
            else:
                evidence_item.document_date = datetime.combine(
                    doc_date_val, datetime.min.time()
                )
        evidence_item.processing_error = None
        evidence_item.processing_status = "processed"
        db.commit()
    except Exception as e:
        evidence_item.processing_status = "error"
        evidence_item.processing_error = str(e)
        evidence_item.extracted_metadata = {
            "extraction_status": "error",
            "extraction_error": str(e),
            "filename": evidence_item.filename,
            "mime_type": evidence_item.mime_type,
        }
        db.commit()
        logger.error(f"Auto-extraction failed for {evidence_item.id}: {e}")
    return {
        "id": str(evidence_item.id),
        "filename": evidence_item.filename,
        "is_duplicate": is_duplicate,
        "duplicate_of_id": str(duplicate_of_id) if duplicate_of_id else None,
        "processing_status": evidence_item.processing_status,
        "message": "Evidence uploaded successfully",
    }


async def list_evidence_service(
    db,
    page,
    page_size,
    search,
    evidence_type,
    document_category,
    date_from,
    date_to,
    tags,
    has_correspondence,
    is_starred,
    is_reviewed,
    include_email_info,
    unassigned,
    case_id,
    project_id,
    collection_id,
    processing_status,
    sort_by,
    sort_order,
    include_hidden,
):
    query = db.query(EvidenceItem)
    if not include_hidden:
        # Filter out evidence items that are directly marked as hidden
        query = query.filter(
            or_(
                EvidenceItem.meta.is_(None),
                EvidenceItem.meta.op("->>")("spam").is_(None),
                EvidenceItem.meta.op("->")("spam").op("->>")("is_hidden") != "true",
            )
        )
        # Also filter out evidence whose source email is hidden/excluded
        # Use outerjoin so evidence without a source email is still included
        query = query.outerjoin(
            EmailMessage, EvidenceItem.source_email_id == EmailMessage.id
        ).filter(
            or_(
                EvidenceItem.source_email_id.is_(None),  # No source email - keep it
                and_(
                    # Source email exists but is not hidden
                    or_(
                        EmailMessage.meta.is_(None),
                        EmailMessage.meta.op("->")("spam")
                        .op("->>")("is_hidden")
                        .is_(None),
                        EmailMessage.meta.op("->")("spam").op("->>")("is_hidden")
                        != "true",
                    ),
                    or_(
                        EmailMessage.meta.is_(None),
                        EmailMessage.meta.op("->")("spam")
                        .op("->>")("user_override")
                        .is_(None),
                        EmailMessage.meta.op("->")("spam").op("->>")("user_override")
                        != "hidden",
                    ),
                    or_(
                        EmailMessage.meta.is_(None),
                        EmailMessage.meta.op("->>")("is_hidden").is_(None),
                        EmailMessage.meta.op("->>")("is_hidden") != "true",
                    ),
                ),
            )
        )
    if search:
        search_term = f"%{search}%"
        query = query.filter(
            or_(
                EvidenceItem.filename.ilike(search_term),
                EvidenceItem.title.ilike(search_term),
                EvidenceItem.extracted_text.ilike(search_term),
                EvidenceItem.description.ilike(search_term),
            )
        )
    if evidence_type:
        query = query.filter(EvidenceItem.evidence_type == evidence_type)
    if document_category:
        query = query.filter(EvidenceItem.document_category == document_category)
    if date_from:
        query = query.filter(EvidenceItem.document_date >= date_from)
    if date_to:
        query = query.filter(EvidenceItem.document_date <= date_to)
    if tags:
        tag_list = [t.strip() for t in tags.split(",")]
        for tag in tag_list:
            query = query.filter(
                or_(
                    EvidenceItem.manual_tags.contains([tag]),
                    EvidenceItem.auto_tags.contains([tag]),
                )
            )
    if is_starred is not None:
        query = query.filter(EvidenceItem.is_starred == is_starred)
    if is_reviewed is not None:
        query = query.filter(EvidenceItem.is_reviewed == is_reviewed)
    if unassigned:
        query = query.filter(
            and_(EvidenceItem.case_id.is_(None), EvidenceItem.project_id.is_(None))
        )
    if case_id:
        try:
            case_uuid = uuid.UUID(case_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid case_id format")
        # If the Case is linked to a Project and the caller did not also request an
        # explicit project_id filter, include project-scoped evidence as well.
        if not project_id:
            linked_project_id = None
            try:
                case = db.query(Case).filter(Case.id == case_uuid).first()
                if case and getattr(case, "project_id", None):
                    linked_project_id = case.project_id
            except Exception:
                linked_project_id = None

            if linked_project_id:
                query = query.filter(
                    or_(
                        EvidenceItem.case_id == case_uuid,
                        EvidenceItem.project_id == linked_project_id,
                    )
                )
            else:
                query = query.filter(EvidenceItem.case_id == case_uuid)
        else:
            query = query.filter(EvidenceItem.case_id == case_uuid)
    if project_id:
        try:
            project_uuid = uuid.UUID(project_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid project_id format")
        query = query.filter(EvidenceItem.project_id == project_uuid)
    if collection_id:
        try:
            collection_uuid = uuid.UUID(collection_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid collection_id format")
        query = query.join(
            EvidenceCollectionItem,
            EvidenceCollectionItem.evidence_item_id == EvidenceItem.id,
        ).filter(EvidenceCollectionItem.collection_id == collection_uuid)
    if processing_status:
        query = query.filter(EvidenceItem.processing_status == processing_status)
    total = query.count()
    sort_column = getattr(EvidenceItem, sort_by, EvidenceItem.created_at)
    if sort_order == "desc":
        query = query.order_by(desc(sort_column))
    else:
        query = query.order_by(sort_column)
    offset = (page - 1) * page_size
    items = query.offset(offset).limit(page_size).all()
    item_ids = [item.id for item in items]
    correspondence_counts = {}
    if item_ids:
        counts = (
            db.query(
                EvidenceCorrespondenceLink.evidence_item_id,
                func.count(EvidenceCorrespondenceLink.id),
            )
            .filter(EvidenceCorrespondenceLink.evidence_item_id.in_(item_ids))
            .group_by(EvidenceCorrespondenceLink.evidence_item_id)
            .all()
        )
        for count_row in counts:
            row_id = count_row[0]
            row_count = count_row[1]
            if row_id is not None:
                count_val = 0
                if isinstance(row_count, int):
                    count_val = row_count
                elif row_count is not None:
                    count_val = int(str(row_count))
                correspondence_counts[str(row_id)] = count_val
    email_info = {}
    source_email_ids = [item.source_email_id for item in items if item.source_email_id]
    if source_email_ids:
        emails = (
            db.query(EmailMessage).filter(EmailMessage.id.in_(source_email_ids)).all()
        )
        email_info = {
            email.id: {
                "subject": email.subject,
                "from": email.sender_email or email.sender_name,
            }
            for email in emails
        }
    summaries = []
    for item in items:
        corr_count = correspondence_counts.get(str(item.id), 0)
        source_email_subject = None
        source_email_sender = None
        if item.source_email_id and item.source_email_id in email_info:
            ei = email_info[item.source_email_id]
            source_email_subject = ei.get("subject")
            source_email_sender = ei.get("from")
        download_url = None
        try:
            download_url = presign_get(item.s3_key, expires=3600, bucket=item.s3_bucket)
        except Exception:
            pass
        summaries.append(
            EvidenceItemSummary(
                id=str(item.id),
                filename=item.filename,
                author=item.author,
                file_type=item.file_type,
                mime_type=item.mime_type,
                file_size=item.file_size,
                page_count=item.page_count,
                evidence_type=item.evidence_type,
                document_category=item.document_category,
                document_date=(
                    item.document_date
                    if isinstance(item.document_date, date)
                    else (item.document_date.date() if item.document_date else None)
                ),
                title=item.title,
                processing_status=item.processing_status or "pending",
                is_starred=item.is_starred or False,
                is_reviewed=item.is_reviewed or False,
                has_correspondence=corr_count > 0,
                correspondence_count=corr_count,
                correspondence_link_count=corr_count,
                auto_tags=item.auto_tags or [],
                manual_tags=item.manual_tags or [],
                extracted_parties=item.extracted_parties or [],
                extracted_amounts=item.extracted_amounts or [],
                case_id=str(item.case_id) if item.case_id else None,
                project_id=str(item.project_id) if item.project_id else None,
                source_type=item.source_type,
                source_email_id=(
                    str(item.source_email_id) if item.source_email_id else None
                ),
                source_email_subject=source_email_subject,
                source_email_sender=source_email_sender,
                download_url=download_url,
                created_at=item.created_at or datetime.now(),
            )
        )
    email_total = 0
    if include_email_info:
        email_query = db.query(EmailMessage)
        if project_id:
            email_query = email_query.filter(
                EmailMessage.project_id == uuid.UUID(project_id)
            )
        if case_id:
            email_query = email_query.filter(EmailMessage.case_id == uuid.UUID(case_id))
        if search:
            search_term = f"%{search}%"
            email_query = email_query.filter(
                or_(
                    EmailMessage.subject.ilike(search_term),
                    EmailMessage.sender_email.ilike(search_term),
                    EmailMessage.body_text.ilike(search_term),
                )
            )
        if date_from:
            email_query = email_query.filter(EmailMessage.date_sent >= date_from)
        if date_to:
            email_query = email_query.filter(EmailMessage.date_sent <= date_to)
        email_total = email_query.count()
        if len(summaries) < page_size:
            evidence_count = total
            email_offset = max(0, (page - 1) * page_size - evidence_count)
            email_limit = page_size - len(summaries)
            if page == 1 or email_offset >= 0:
                emails = (
                    email_query.order_by(desc(EmailMessage.date_sent))
                    .offset(email_offset)
                    .limit(email_limit)
                    .all()
                )
                for email in emails:
                    summaries.append(
                        EvidenceItemSummary(
                            id=f"email-{email.id}",
                            filename=f"{email.subject or 'No Subject'}.eml",
                            file_type="eml",
                            mime_type="message/rfc822",
                            file_size=len(email.body_text or "")
                            + len(email.body_html or ""),
                            evidence_type="correspondence",
                            document_category="email",
                            document_date=(
                                email.date_sent.date() if email.date_sent else None
                            ),
                            title=email.subject,
                            processing_status="completed",
                            is_starred=False,
                            is_reviewed=False,
                            has_correspondence=True,
                            correspondence_count=1,
                            correspondence_link_count=0,
                            auto_tags=[],
                            manual_tags=[],
                            extracted_parties=[],
                            extracted_amounts=[],
                            case_id=str(email.case_id) if email.case_id else None,
                            project_id=(
                                str(email.project_id) if email.project_id else None
                            ),
                            source_type="pst",
                            source_email_id=str(email.id),
                            source_email_subject=email.subject,
                            source_email_sender=email.sender_email,
                            download_url=None,
                            created_at=email.created_at or datetime.now(),
                        )
                    )
    return EvidenceListResponse(
        total=total + email_total, items=summaries, page=page, page_size=page_size
    )


async def get_evidence_server_side_service(
    request, db, project_id, case_id, collection_id, include_email_info, include_hidden
):
    project_uuid = None
    case_uuid = None
    collection_uuid = None
    if project_id:
        try:
            project_uuid = uuid.UUID(project_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid project_id format")
    if case_id:
        try:
            case_uuid = uuid.UUID(case_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid case_id format")
    if collection_id:
        try:
            collection_uuid = uuid.UUID(collection_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid collection_id format")
    base_query = db.query(EvidenceItem)
    if not include_hidden:
        # Filter out evidence items that are directly marked as hidden
        base_query = base_query.filter(
            or_(
                EvidenceItem.meta.is_(None),
                EvidenceItem.meta.op("->>")("spam").is_(None),
                EvidenceItem.meta.op("->")("spam").op("->>")("is_hidden") != "true",
            )
        )
        # Also filter out evidence whose source email is hidden/excluded
        base_query = base_query.outerjoin(
            EmailMessage, EvidenceItem.source_email_id == EmailMessage.id
        ).filter(
            or_(
                EvidenceItem.source_email_id.is_(None),  # No source email - keep it
                and_(
                    # Source email exists but is not hidden
                    or_(
                        EmailMessage.meta.is_(None),
                        EmailMessage.meta.op("->")("spam")
                        .op("->>")("is_hidden")
                        .is_(None),
                        EmailMessage.meta.op("->")("spam").op("->>")("is_hidden")
                        != "true",
                    ),
                    or_(
                        EmailMessage.meta.is_(None),
                        EmailMessage.meta.op("->")("spam")
                        .op("->>")("user_override")
                        .is_(None),
                        EmailMessage.meta.op("->")("spam").op("->>")("user_override")
                        != "hidden",
                    ),
                    or_(
                        EmailMessage.meta.is_(None),
                        EmailMessage.meta.op("->>")("is_hidden").is_(None),
                        EmailMessage.meta.op("->>")("is_hidden") != "true",
                    ),
                ),
            )
        )
    if project_uuid is not None:
        base_query = base_query.filter(EvidenceItem.project_id == project_uuid)
    if case_uuid is not None:
        base_query = base_query.filter(EvidenceItem.case_id == case_uuid)
    if collection_uuid is not None:
        base_query = base_query.join(
            EvidenceCollectionItem,
            EvidenceCollectionItem.evidence_item_id == EvidenceItem.id,
        ).filter(EvidenceCollectionItem.collection_id == collection_uuid)
    filtered_query = _apply_ag_filters(base_query, request.filterModel)
    sorted_query = _apply_ag_sorting(filtered_query, request.sortModel)

    start_row = max(0, int(getattr(request, "startRow", 0) or 0))
    end_row = max(start_row, int(getattr(request, "endRow", 0) or (start_row + 100)))
    block_size = max(end_row - start_row, 0) or 100

    page_query = sorted_query.offset(start_row).limit(block_size)
    items = page_query.all()
    # lastRow semantics:
    # - return -1 when we don't know the total yet (more rows likely exist)
    # - return the exact row count only when we've reached the end of the dataset
    last_row = (start_row + len(items)) if len(items) < block_size else -1
    item_ids = [item.id for item in items]
    source_email_ids = [item.source_email_id for item in items if item.source_email_id]
    correspondence_counts = {}
    if item_ids:
        counts = (
            db.query(
                EvidenceCorrespondenceLink.evidence_item_id,
                func.count(EvidenceCorrespondenceLink.id),
            )
            .filter(EvidenceCorrespondenceLink.evidence_item_id.in_(item_ids))
            .group_by(EvidenceCorrespondenceLink.evidence_item_id)
            .all()
        )
        for count_row in counts:
            row_id = count_row[0]
            row_count = count_row[1]
            if row_id is not None:
                count_val = 0
                if isinstance(row_count, int):
                    count_val = row_count
                elif row_count is not None:
                    count_val = int(str(row_count))
                correspondence_counts[str(row_id)] = count_val
    email_info: dict[uuid.UUID, dict[str, Any]] = {}
    if source_email_ids:
        emails = (
            db.query(EmailMessage).filter(EmailMessage.id.in_(source_email_ids)).all()
        )
        email_info = {}
        for email in emails:
            email_info[email.id] = {
                "subject": email.subject,
                "sender": email.sender_email or email.sender_name,
                "sender_email": email.sender_email,
                "sender_name": email.sender_name,
                "recipients_to": email.recipients_to or [],
                "recipients_cc": email.recipients_cc or [],
                "date": _extract_email_date(email),
                "preview": _extract_email_preview(email),
                "message_id": email.message_id,
                "references": email.email_references,
            }
    stakeholder_domain_map: dict[str, str] = {}
    stakeholder_email_map: dict[str, str] = {}
    if source_email_ids and (project_uuid or case_uuid):
        stakeholder_domain_map, stakeholder_email_map = _load_stakeholder_maps(
            db, project_uuid, case_uuid
        )
    rows = []
    for item in items:
        corr_count = correspondence_counts.get(str(item.id), 0)
        email_meta = (
            email_info.get(item.source_email_id) if item.source_email_id else None
        )
        email_date = email_meta.get("date") if email_meta else None
        metadata_date = None
        if item.extracted_metadata and not item.document_date and not email_date:
            metadata_date = _extract_metadata_date(item.extracted_metadata)
        filename_date = None
        if not item.document_date and not email_date and not metadata_date:
            filename_date = _parse_filename_date(item.filename)
        parties = item.extracted_parties or []
        if not parties and email_meta:
            parties = _extract_company_parties(
                email_meta.get("sender_email"),
                email_meta.get("recipients_to"),
                email_meta.get("recipients_cc"),
                stakeholder_domain_map,
                stakeholder_email_map,
            )
        fallback_date = email_date or metadata_date or filename_date
        primary_date = item.document_date
        if primary_date and item.created_at:
            created_date = _resolve_document_date(item.created_at, None)
            primary_date_val = _resolve_document_date(primary_date, None)
            source_type = (item.source_type or "").lower()
            if (
                created_date
                and primary_date_val == created_date
                and (
                    item.source_email_id
                    or source_type
                    in {
                        "pst_extraction",
                        "pst",
                        "email",
                        "email_attachment",
                        "email-attachment",
                    }
                )
            ):
                primary_date = None
        document_date = _resolve_document_date(primary_date, fallback_date)
        rows.append(
            {
                "id": str(item.id),
                "filename": item.filename,
                "title": item.title,
                "mime_type": item.mime_type,
                "file_type": item.file_type,
                "file_size": item.file_size,
                "evidence_type": item.evidence_type,
                "document_category": item.document_category,
                "document_date": document_date,
                "processing_status": item.processing_status or "pending",
                "is_starred": item.is_starred or False,
                "is_reviewed": item.is_reviewed or False,
                "has_correspondence": corr_count > 0,
                "correspondence_count": corr_count,
                "correspondence_link_count": corr_count,
                "auto_tags": item.auto_tags or [],
                "manual_tags": item.manual_tags or [],
                "extracted_parties": parties,
                "extracted_amounts": item.extracted_amounts or [],
                "author": item.author,
                "page_count": item.page_count,
                "case_id": str(item.case_id) if item.case_id else None,
                "project_id": str(item.project_id) if item.project_id else None,
                "source_type": item.source_type,
                "source_email_id": (
                    str(item.source_email_id) if item.source_email_id else None
                ),
                "source_email_subject": (
                    email_meta.get("subject") if email_meta else None
                ),
                "source_email_sender": (
                    email_meta.get("sender") if email_meta else None
                ),
                "source_email_date": (email_meta.get("date") if email_meta else None),
                "source_email_preview": (
                    email_meta.get("preview") if email_meta else None
                ),
                "source_email_message_id": (
                    email_meta.get("message_id") if email_meta else None
                ),
                "source_email_references": (
                    email_meta.get("references") if email_meta else None
                ),
                "source_email_to": (
                    email_meta.get("recipients_to") if email_meta else None
                ),
                "source_email_cc": (
                    email_meta.get("recipients_cc") if email_meta else None
                ),
                "created_at": item.created_at or datetime.now(),
            }
        )
    # Calculate stats for UI tab counts
    # Use base_query (before AG-Grid filtering) for accurate total counts
    scope_query = db.query(EvidenceItem)
    if not include_hidden:
        scope_query = scope_query.filter(
            or_(
                EvidenceItem.meta.is_(None),
                EvidenceItem.meta.op("->>")("spam").is_(None),
                EvidenceItem.meta.op("->")("spam").op("->>")("is_hidden") != "true",
            )
        )
    if project_uuid is not None:
        scope_query = scope_query.filter(EvidenceItem.project_id == project_uuid)
    if case_uuid is not None:
        scope_query = scope_query.filter(EvidenceItem.case_id == case_uuid)

    total_count = scope_query.count()

    image_exts = [
        "jpg",
        "jpeg",
        "png",
        "gif",
        "bmp",
        "webp",
        "tif",
        "tiff",
        "heic",
        "heif",
        "svg",
    ]
    image_types = ["image", "photo", "photograph"]
    image_count = scope_query.filter(
        or_(
            EvidenceItem.mime_type.ilike("image/%"),
            EvidenceItem.file_type.in_(image_exts),
            EvidenceItem.evidence_type.in_(image_types),
        )
    ).count()

    stats = {
        "total": total_count,
        "image_count": image_count,
    }
    return ServerSideEvidenceResponse(rows=rows, lastRow=last_row, stats=stats)


async def get_evidence_download_url_service(evidence_id, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    try:
        url = presign_get(
            item.s3_key,
            expires=3600,
            bucket=item.s3_bucket,
            response_disposition=f'attachment; filename="{item.filename}"',
        )
    except Exception as e:
        logger.error(f"Failed to generate download URL for {evidence_id}: {e}")
        raise HTTPException(500, "Unable to generate download URL")
    return {"evidence_id": evidence_id, "download_url": url}


async def get_evidence_full_service(evidence_id, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")

    # Debug logging for S3 path troubleshooting
    logger.info(
        f"[EVIDENCE DEBUG] Loading item {evidence_id}: "
        f"s3_bucket={item.s3_bucket!r}, s3_key={item.s3_key!r}, filename={item.filename!r}"
    )

    download_url = None
    try:
        download_url = presign_get(
            item.s3_key,
            expires=3600,
            bucket=item.s3_bucket,
            response_disposition=f'attachment; filename="{item.filename}"',
        )
    except Exception as e:
        logger.warning(
            f"Could not create download URL: {e} - bucket={item.s3_bucket}, key={item.s3_key}"
        )
    preview_type = "unsupported"
    preview_url = None
    preview_content = None
    page_count = None
    dimensions = None
    mime_type = item.mime_type or "application/octet-stream"
    try:
        preview_url = presign_get(
            item.s3_key,
            expires=3600,
            bucket=item.s3_bucket,
            response_content_type=mime_type,
        )
    except Exception as e:
        logger.warning(f"Could not generate preview URL: {e}")
    if mime_type.startswith("image/"):
        preview_type = "image"
        if item.extracted_metadata:
            dimensions = {
                "width": item.extracted_metadata.get("width"),
                "height": item.extracted_metadata.get("height"),
            }
    elif mime_type == "application/pdf":
        preview_type = "pdf"
        if item.extracted_metadata:
            page_count = item.extracted_metadata.get("page_count")
    elif mime_type.startswith("text/") or item.filename.lower().endswith(
        (".txt", ".csv", ".json", ".xml", ".html", ".md", ".log")
    ):
        preview_type = "text"
        preview_content = "Text preview available via /text-content"
    elif mime_type.startswith("audio/"):
        preview_type = "audio"
    elif mime_type.startswith("video/"):
        preview_type = "video"
    elif mime_type in [
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-powerpoint",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ]:
        preview_type = "office"
        if item.extracted_metadata:
            preview_content = item.extracted_metadata.get("text_preview")
            page_count = item.extracted_metadata.get(
                "page_count"
            ) or item.extracted_metadata.get("slide_count")
    email_meta = None
    if item.source_email_id:
        email = (
            db.query(EmailMessage)
            .filter(EmailMessage.id == item.source_email_id)
            .first()
        )
        if email:
            email_meta = {
                "subject": email.subject,
                "sender": email.sender_email or email.sender_name,
                "date": _extract_email_date(email),
                "preview": _extract_email_preview(email),
                "sender_email": email.sender_email,
                "recipients_to": email.recipients_to or [],
                "recipients_cc": email.recipients_cc or [],
                "message_id": email.message_id,
                "references": email.email_references,
                "project_id": email.project_id,
                "case_id": email.case_id,
            }
    metadata_date = None
    if (
        item.extracted_metadata
        and not item.document_date
        and not (email_meta and email_meta.get("date"))
    ):
        metadata_date = _extract_metadata_date(item.extracted_metadata)
    parties = item.extracted_parties or []
    if not parties and email_meta:
        domain_map, email_map = _load_stakeholder_maps(
            db, email_meta.get("project_id"), email_meta.get("case_id")
        )
        parties = _extract_company_parties(
            email_meta.get("sender_email"),
            email_meta.get("recipients_to"),
            email_meta.get("recipients_cc"),
            domain_map,
            email_map,
        )
    fallback_date = (
        (email_meta.get("date") if email_meta else None)
        or metadata_date
        or _parse_filename_date(item.filename)
    )
    primary_date = item.document_date
    if primary_date and item.created_at:
        created_date = _resolve_document_date(item.created_at, None)
        primary_date_val = _resolve_document_date(primary_date, None)
        source_type = (item.source_type or "").lower()
        if (
            created_date
            and primary_date_val == created_date
            and (
                item.source_email_id
                or source_type
                in {
                    "pst_extraction",
                    "pst",
                    "email",
                    "email_attachment",
                    "email-attachment",
                }
            )
        ):
            primary_date = None
    document_date = _resolve_document_date(primary_date, fallback_date)
    detail = {
        "id": str(item.id),
        "filename": item.filename,
        "title": item.title,
        "author": item.author,
        "file_type": item.file_type,
        "mime_type": item.mime_type,
        "file_size": item.file_size,
        "file_hash": item.file_hash,
        "evidence_type": item.evidence_type,
        "document_category": item.document_category,
        "document_date": document_date,
        "processing_status": item.processing_status or "pending",
        "page_count": item.page_count,
        "auto_tags": item.auto_tags or [],
        "manual_tags": item.manual_tags or [],
        "extracted_parties": _safe_dict_list(parties),
        "extracted_dates": _safe_dict_list(item.extracted_dates),
        "extracted_amounts": _safe_dict_list(item.extracted_amounts),
        "extracted_references": _safe_dict_list(item.extracted_references),
        "source_type": item.source_type,
        "source_email_id": str(item.source_email_id) if item.source_email_id else None,
        "source_email_subject": email_meta.get("subject") if email_meta else None,
        "source_email_sender": email_meta.get("sender") if email_meta else None,
        "source_email_date": email_meta.get("date") if email_meta else None,
        "source_email_preview": email_meta.get("preview") if email_meta else None,
        "source_email_message_id": (
            email_meta.get("message_id") if email_meta else None
        ),
        "source_email_references": (
            email_meta.get("references") if email_meta else None
        ),
        "source_email_to": email_meta.get("recipients_to") if email_meta else None,
        "source_email_cc": email_meta.get("recipients_cc") if email_meta else None,
        "case_id": str(item.case_id) if item.case_id else None,
        "project_id": str(item.project_id) if item.project_id else None,
        "collection_id": str(item.collection_id) if item.collection_id else None,
        "is_starred": item.is_starred or False,
        "is_privileged": item.is_privileged or False,
        "is_confidential": item.is_confidential or False,
        "is_reviewed": item.is_reviewed or False,
        "notes": item.notes,
        "created_at": item.created_at or datetime.now(),
        "updated_at": item.updated_at,
    }
    preview_payload = {
        "preview_type": preview_type,
        "preview_url": preview_url,
        "preview_content": preview_content,
        "page_count": page_count,
        "dimensions": dimensions,
        "download_url": download_url,
        "filename": item.filename,
        "mime_type": mime_type,
        # Debug info - shows S3 location for troubleshooting
        "_debug_s3_bucket": item.s3_bucket,
        "_debug_s3_key": item.s3_key,
    }
    metadata = item.extracted_metadata or {}
    return {
        "detail": detail,
        "preview": preview_payload,
        "metadata": metadata,
        "download_url": download_url,
    }


async def get_evidence_detail_service(evidence_id, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    links = (
        db.query(EvidenceCorrespondenceLink)
        .filter(EvidenceCorrespondenceLink.evidence_item_id == evidence_uuid)
        .all()
    )
    correspondence_links = []
    for link in links:
        link_data = {
            "id": str(link.id),
            "link_type": link.link_type,
            "link_confidence": link.link_confidence,
            "is_verified": link.is_verified,
            "context_snippet": link.context_snippet,
        }
        if link.email_message_id:
            email = (
                db.query(EmailMessage)
                .filter(EmailMessage.id == link.email_message_id)
                .first()
            )
            if email:
                link_data["email"] = {
                    "id": str(email.id),
                    "subject": email.subject,
                    "sender": email.sender_email or email.sender_name,
                    "date": email.date_sent.isoformat() if email.date_sent else None,
                }
        else:
            link_data["correspondence"] = {
                "type": link.correspondence_type,
                "reference": link.correspondence_reference,
                "date": (
                    link.correspondence_date.isoformat()
                    if link.correspondence_date
                    else None
                ),
                "from": link.correspondence_from,
                "to": link.correspondence_to,
                "subject": link.correspondence_subject,
            }
        correspondence_links.append(link_data)
    relations = (
        db.query(EvidenceRelation)
        .filter(
            or_(
                EvidenceRelation.source_evidence_id == evidence_uuid,
                EvidenceRelation.target_evidence_id == evidence_uuid,
            )
        )
        .all()
    )
    relation_list = []
    for rel in relations:
        other_id = (
            rel.target_evidence_id
            if rel.source_evidence_id == evidence_uuid
            else rel.source_evidence_id
        )
        other_item = db.query(EvidenceItem).filter(EvidenceItem.id == other_id).first()
        rel_data = {
            "id": str(rel.id),
            "relation_type": rel.relation_type,
            "direction": (
                "outgoing" if rel.source_evidence_id == evidence_uuid else "incoming"
            ),
            "is_verified": rel.is_verified,
            "related_item": (
                {
                    "id": str(other_item.id),
                    "filename": other_item.filename,
                    "title": other_item.title,
                }
                if other_item
                else None
            ),
        }
        relation_list.append(rel_data)
    download_url = presign_get(
        item.s3_key,
        expires=3600,
        bucket=item.s3_bucket,
        response_disposition=f'attachment; filename="{item.filename}"',
    )
    user = get_default_user(db)
    log_activity(db, "view", user.id, evidence_item_id=item.id)
    db.commit()
    return EvidenceItemDetail(
        id=str(item.id),
        filename=item.filename,
        original_path=item.original_path,
        file_type=item.file_type,
        mime_type=item.mime_type,
        file_size=item.file_size,
        file_hash=item.file_hash,
        evidence_type=item.evidence_type,
        document_category=item.document_category,
        document_date=(
            item.document_date
            if isinstance(item.document_date, date)
            else (item.document_date.date() if item.document_date else None)
        ),
        title=item.title,
        author=item.author,
        description=item.description,
        page_count=item.page_count,
        extracted_text=item.extracted_text[:5000] if item.extracted_text else None,
        extracted_parties=_safe_dict_list(item.extracted_parties),
        extracted_dates=_safe_dict_list(item.extracted_dates),
        extracted_amounts=_safe_dict_list(item.extracted_amounts),
        extracted_references=_safe_dict_list(item.extracted_references),
        auto_tags=item.auto_tags or [],
        manual_tags=item.manual_tags or [],
        processing_status=item.processing_status or "pending",
        source_type=item.source_type,
        source_path=item.source_path,
        is_duplicate=item.is_duplicate or False,
        is_starred=item.is_starred or False,
        is_privileged=item.is_privileged or False,
        is_confidential=item.is_confidential or False,
        is_reviewed=item.is_reviewed or False,
        notes=item.notes,
        case_id=str(item.case_id) if item.case_id else None,
        project_id=str(item.project_id) if item.project_id else None,
        collection_id=str(item.collection_id) if item.collection_id else None,
        correspondence_links=correspondence_links,
        relations=relation_list,
        download_url=download_url,
        created_at=item.created_at or datetime.now(),
        updated_at=item.updated_at,
    )


async def update_evidence_service(evidence_id, updates, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    update_data = updates.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        typed_value = value
        if field in ["case_id", "project_id", "collection_id"] and typed_value:
            typed_value = uuid.UUID(str(typed_value))
        setattr(item, field, typed_value)
    user = get_default_user(db)
    log_activity(
        db,
        "update",
        user.id,
        evidence_item_id=item.id,
        details={"updated_fields": list(update_data.keys())},
    )
    db.commit()
    db.refresh(item)
    return {"id": str(item.id), "message": "Evidence updated successfully"}


async def delete_evidence_service(evidence_id, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    user = get_default_user(db)
    log_activity(
        db,
        "delete",
        user.id,
        details={"evidence_id": evidence_id, "filename": item.filename},
    )
    try:
        s3_client = s3()
        if s3_client is not None:
            s3_client.delete_object(Bucket=item.s3_bucket, Key=item.s3_key)
    except Exception as e:
        logger.warning(f"Failed to delete S3 object: {e}")
    db.delete(item)
    db.commit()
    return {"message": "Evidence deleted successfully"}


async def assign_evidence_service(evidence_id, assignment, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    if assignment.case_id:
        case = db.query(Case).filter(Case.id == uuid.UUID(assignment.case_id)).first()
        if not case:
            raise HTTPException(404, "Case not found")
        item.case_id = case.id
    if assignment.project_id:
        project = (
            db.query(Project)
            .filter(Project.id == uuid.UUID(assignment.project_id))
            .first()
        )
        if not project:
            raise HTTPException(404, "Project not found")
        item.project_id = project.id
    user = get_default_user(db)
    log_activity(
        db,
        "assign",
        user.id,
        evidence_item_id=item.id,
        details={"case_id": assignment.case_id, "project_id": assignment.project_id},
    )
    db.commit()
    return {
        "id": str(item.id),
        "case_id": str(item.case_id) if item.case_id else None,
        "project_id": str(item.project_id) if item.project_id else None,
        "message": "Evidence assigned successfully",
    }


async def toggle_star_service(evidence_id, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    item.is_starred = not item.is_starred
    db.commit()
    return {"id": str(item.id), "is_starred": item.is_starred}


async def get_evidence_correspondence_service(evidence_id, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    links = (
        db.query(EvidenceCorrespondenceLink)
        .filter(EvidenceCorrespondenceLink.evidence_item_id == evidence_uuid)
        .all()
    )
    result = []
    for link in links:
        link_data = {
            "id": str(link.id),
            "link_type": link.link_type,
            "link_confidence": link.link_confidence,
            "link_method": link.link_method,
            "is_auto_linked": link.is_auto_linked,
            "is_verified": link.is_verified,
            "context_snippet": link.context_snippet,
            "page_reference": link.page_reference,
            "created_at": link.created_at.isoformat() if link.created_at else None,
        }
        if link.email_message_id:
            email = (
                db.query(EmailMessage)
                .filter(EmailMessage.id == link.email_message_id)
                .first()
            )
            if email:
                link_data["email"] = {
                    "id": str(email.id),
                    "subject": email.subject,
                    "sender_email": email.sender_email,
                    "sender_name": email.sender_name,
                    "date_sent": (
                        email.date_sent.isoformat() if email.date_sent else None
                    ),
                    "has_attachments": email.has_attachments,
                }
        else:
            link_data["external_correspondence"] = {
                "type": link.correspondence_type,
                "reference": link.correspondence_reference,
                "date": (
                    link.correspondence_date.isoformat()
                    if link.correspondence_date
                    else None
                ),
                "from": link.correspondence_from,
                "to": link.correspondence_to,
                "subject": link.correspondence_subject,
            }
        result.append(link_data)
    return {"evidence_id": evidence_id, "links": result, "total": len(result)}


async def link_evidence_to_email_service(evidence_id, link_request, db):
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid evidence ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    email_message_id = None
    if link_request.email_message_id:
        email = (
            db.query(EmailMessage)
            .filter(EmailMessage.id == uuid.UUID(link_request.email_message_id))
            .first()
        )
        if not email:
            raise HTTPException(404, "Email message not found")
        email_message_id = email.id
        existing = (
            db.query(EvidenceCorrespondenceLink)
            .filter(
                and_(
                    EvidenceCorrespondenceLink.evidence_item_id == evidence_uuid,
                    EvidenceCorrespondenceLink.email_message_id == email_message_id,
                )
            )
            .first()
        )
        if existing:
            raise HTTPException(409, "Link already exists")
    user = get_default_user(db)
    link = EvidenceCorrespondenceLink(
        evidence_item_id=evidence_uuid,
        email_message_id=email_message_id,
        link_type=link_request.link_type,
        link_confidence=100,
        link_method="manual",
        correspondence_type=link_request.correspondence_type,
        correspondence_reference=link_request.correspondence_reference,
        correspondence_date=link_request.correspondence_date,
        correspondence_from=link_request.correspondence_from,
        correspondence_to=link_request.correspondence_to,
        correspondence_subject=link_request.correspondence_subject,
        context_snippet=link_request.context_snippet,
        is_auto_linked=False,
        is_verified=True,
        linked_by=user.id,
        verified_by=user.id,
        verified_at=datetime.now(),
    )
    db.add(link)
    log_activity(
        db,
        "link",
        user.id,
        evidence_item_id=evidence_uuid,
        details={
            "email_id": link_request.email_message_id,
            "link_type": link_request.link_type,
        },
    )
    db.commit()
    db.refresh(link)
    return {"id": str(link.id), "message": "Link created successfully"}


async def delete_correspondence_link_service(link_id, db):
    try:
        link_uuid = uuid.UUID(link_id)
    except ValueError:
        raise HTTPException(400, "Invalid link ID format")
    link = (
        db.query(EvidenceCorrespondenceLink)
        .filter(EvidenceCorrespondenceLink.id == link_uuid)
        .first()
    )
    if not link:
        raise HTTPException(404, "Link not found")
    db.delete(link)
    db.commit()
    return {"message": "Link deleted successfully"}


async def list_collections_service(db, include_system, case_id, project_id):
    query = db.query(EvidenceCollection)
    if not include_system:
        query = query.filter(EvidenceCollection.is_system == False)
    if case_id:
        try:
            case_uuid = uuid.UUID(case_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid case_id format")
        query = query.filter(
            or_(
                EvidenceCollection.case_id == case_uuid,
                EvidenceCollection.case_id.is_(None),
            )
        )
    if project_id:
        try:
            project_uuid = uuid.UUID(project_id)
        except ValueError:
            raise HTTPException(status_code=400, detail="Invalid project_id format")
        query = query.filter(
            or_(
                EvidenceCollection.project_id == project_uuid,
                EvidenceCollection.project_id.is_(None),
            )
        )
    collections = query.order_by(
        EvidenceCollection.sort_order, EvidenceCollection.name
    ).all()

    # Always compute item counts from the junction table so the UI reflects reality.
    # The `evidence_collections.item_count` column is not reliably maintained.
    counts_by_collection: dict[uuid.UUID, int] = {}
    collection_ids = [c.id for c in collections]
    if collection_ids:
        count_rows = (
            db.query(
                EvidenceCollectionItem.collection_id,
                func.count(EvidenceCollectionItem.id),
            )
            .filter(EvidenceCollectionItem.collection_id.in_(collection_ids))
            .group_by(EvidenceCollectionItem.collection_id)
            .all()
        )
        for cid, cnt in count_rows:
            if cid is None:
                continue
            counts_by_collection[cid] = int(cnt or 0)

    return [
        CollectionSummary(
            id=str(c.id),
            name=c.name,
            description=c.description,
            collection_type=c.collection_type or "manual",
            parent_id=str(c.parent_id) if c.parent_id else None,
            item_count=counts_by_collection.get(c.id, 0),
            is_system=c.is_system or False,
            color=c.color,
            icon=c.icon,
            case_id=str(c.case_id) if c.case_id else None,
            project_id=str(c.project_id) if c.project_id else None,
        )
        for c in collections
    ]


async def auto_categorize_evidence_service(
    request: AutoCategorizeRequest,
    db,
    project_id: str | None = None,
    case_id: str | None = None,
    include_hidden: bool = False,
) -> AutoCategorizeResponse:
    """Server-side auto-categorization.

    This avoids hundreds of per-item PATCH requests from the browser and keeps
    categorization consistent across clients.
    """

    project_uuid: uuid.UUID | None = None
    case_uuid: uuid.UUID | None = None
    if project_id:
        try:
            project_uuid = uuid.UUID(str(project_id))
        except ValueError as exc:
            raise HTTPException(
                status_code=400, detail="Invalid project_id format"
            ) from exc
    if case_id:
        try:
            case_uuid = uuid.UUID(str(case_id))
        except ValueError as exc:
            raise HTTPException(
                status_code=400, detail="Invalid case_id format"
            ) from exc

    # In case context, derive the linked project so we can also use/create
    # project-level collections consistently.
    if case_uuid is not None and project_uuid is None:
        try:
            case = db.query(Case).filter(Case.id == case_uuid).first()
            if case and getattr(case, "project_id", None):
                project_uuid = case.project_id
        except Exception:
            # Keep None if anything goes wrong; case-scoped categorization still works.
            project_uuid = None

    q = db.query(EvidenceItem)
    if not include_hidden:
        q = q.filter(
            or_(
                EvidenceItem.meta.is_(None),
                EvidenceItem.meta.op("->>")("spam").is_(None),
                EvidenceItem.meta.op("->")("spam").op("->>")("is_hidden") != "true",
            )
        )
    if project_uuid is not None:
        q = q.filter(EvidenceItem.project_id == project_uuid)
    if case_uuid is not None:
        q = q.filter(EvidenceItem.case_id == case_uuid)

    q = q.filter(
        or_(
            EvidenceItem.document_category.is_(None),
            EvidenceItem.document_category == "",
        )
    )

    max_items = int(request.max_items or 0)
    if max_items <= 0:
        max_items = 500
    max_items = min(max_items, 5000)  # hard cap for safety

    items = q.order_by(desc(EvidenceItem.created_at)).limit(max_items).all()

    # Infer categories
    inferred: list[tuple[EvidenceItem, str]] = []
    category_counts: dict[str, int] = {}
    for item in items:
        cat = infer_document_category(
            filename=getattr(item, "filename", None),
            title=getattr(item, "title", None),
            evidence_type=getattr(item, "evidence_type", None),
        )
        if not cat:
            continue
        inferred.append((item, cat))
        category_counts[cat] = category_counts.get(cat, 0) + 1

    if not inferred:
        return AutoCategorizeResponse(
            categorized=0,
            created_collections=0,
            added_to_collections=0,
            categories={},
        )

    created_collections = 0
    added_to_collections = 0

    # Prepare collections (by category name)
    category_to_collection_id: dict[str, uuid.UUID] = {}
    if request.create_collections or request.add_to_collections:
        lower_names = {c.lower() for c in category_counts.keys()}
        existing_query = (
            db.query(EvidenceCollection)
            .filter(EvidenceCollection.is_system.is_(False))
            .filter(func.lower(EvidenceCollection.name).in_(list(lower_names)))
        )

        # Scope existing collection lookup to the current context.
        if case_uuid is not None:
            scope_parts: list[Any] = [EvidenceCollection.case_id == case_uuid]
            if project_uuid is not None:
                scope_parts.append(
                    and_(
                        EvidenceCollection.project_id == project_uuid,
                        EvidenceCollection.case_id.is_(None),
                    )
                )
            existing_query = existing_query.filter(or_(*scope_parts))
        elif project_uuid is not None:
            existing_query = existing_query.filter(
                and_(
                    EvidenceCollection.project_id == project_uuid,
                    EvidenceCollection.case_id.is_(None),
                )
            )

        existing = existing_query.all()
        for col in existing:
            category_to_collection_id[(col.name or "").lower()] = col.id

        if request.create_collections and not request.dry_run:
            user = get_default_user(db)
            for cat in category_counts.keys():
                key = cat.lower()
                if key in category_to_collection_id:
                    continue
                new_collection = EvidenceCollection(
                    name=cat,
                    description=f"Auto-created collection for {cat}",
                    collection_type="manual",
                    filter_rules={},
                    parent_id=None,
                    path=f"/{cat}",
                    depth=0,
                    case_id=case_uuid,
                    project_id=project_uuid,
                    color=None,
                    icon=None,
                    is_system=False,
                    created_by=user.id,
                )
                db.add(new_collection)
                db.flush()  # allocate id
                category_to_collection_id[key] = new_collection.id
                created_collections += 1

    # Add collection memberships (bulk-ish)
    if request.add_to_collections and category_to_collection_id and not request.dry_run:
        collection_ids = list(category_to_collection_id.values())
        evidence_ids = [it.id for it, _cat in inferred]
        existing_pairs: set[tuple[uuid.UUID, uuid.UUID]] = set()
        if collection_ids and evidence_ids:
            pair_rows = (
                db.query(
                    EvidenceCollectionItem.collection_id,
                    EvidenceCollectionItem.evidence_item_id,
                )
                .filter(
                    EvidenceCollectionItem.collection_id.in_(collection_ids),
                    EvidenceCollectionItem.evidence_item_id.in_(evidence_ids),
                )
                .all()
            )
            existing_pairs = {
                (row[0], row[1])
                for row in pair_rows
                if row[0] is not None and row[1] is not None
            }

        user = get_default_user(db)
        for item, cat in inferred:
            col_id = category_to_collection_id.get(cat.lower())
            if not col_id:
                continue
            key = (col_id, item.id)
            if key in existing_pairs:
                continue
            db.add(
                EvidenceCollectionItem(
                    collection_id=col_id,
                    evidence_item_id=item.id,
                    added_method="auto",
                    added_by=user.id,
                )
            )
            existing_pairs.add(key)
            added_to_collections += 1

    # Persist category updates
    categorized = 0
    if not request.dry_run:
        for item, cat in inferred:
            item.document_category = cat
            categorized += 1
        user = get_default_user(db)
        log_activity(
            db,
            "auto_categorize",
            user.id,
            details={
                "categorized": categorized,
                "created_collections": created_collections,
                "added_to_collections": added_to_collections,
                "categories": category_counts,
            },
        )
        db.commit()
    else:
        categorized = len(inferred)

    return AutoCategorizeResponse(
        categorized=categorized,
        created_collections=created_collections,
        added_to_collections=added_to_collections,
        categories=category_counts,
    )


async def create_collection_service(collection, db):
    user = get_default_user(db)
    path = f"/{collection.name}"
    depth = 0
    if collection.parent_id:
        parent = (
            db.query(EvidenceCollection)
            .filter(EvidenceCollection.id == uuid.UUID(collection.parent_id))
            .first()
        )
        if parent:
            path = f"{parent.path}/{collection.name}"
            depth = (parent.depth or 0) + 1
    new_collection = EvidenceCollection(
        name=collection.name,
        description=collection.description,
        collection_type="manual" if not collection.filter_rules else "smart",
        filter_rules=collection.filter_rules or {},
        parent_id=uuid.UUID(collection.parent_id) if collection.parent_id else None,
        path=path,
        depth=depth,
        case_id=uuid.UUID(collection.case_id) if collection.case_id else None,
        project_id=uuid.UUID(collection.project_id) if collection.project_id else None,
        color=collection.color,
        icon=collection.icon,
        is_system=False,
        created_by=user.id,
    )
    db.add(new_collection)
    log_activity(
        db,
        "create_collection",
        user.id,
        collection_id=new_collection.id,
        details={"name": collection.name},
    )
    db.commit()
    db.refresh(new_collection)
    return {
        "id": str(new_collection.id),
        "name": new_collection.name,
        "message": "Collection created successfully",
    }


async def update_collection_service(collection_id, updates, db):
    try:
        collection_uuid = uuid.UUID(collection_id)
    except ValueError:
        raise HTTPException(400, "Invalid collection ID format")
    collection = (
        db.query(EvidenceCollection)
        .filter(EvidenceCollection.id == collection_uuid)
        .first()
    )
    if not collection:
        raise HTTPException(404, "Collection not found")
    if collection.is_system:
        raise HTTPException(403, "Cannot modify system collections")
    update_data = updates.model_dump(exclude_unset=True)
    for field, value in update_data.items():
        setattr(collection, field, value)
    db.commit()
    db.refresh(collection)
    return {"id": str(collection.id), "message": "Collection updated successfully"}


async def delete_collection_service(collection_id, db):
    try:
        collection_uuid = uuid.UUID(collection_id)
    except ValueError:
        raise HTTPException(400, "Invalid collection ID format")
    collection = (
        db.query(EvidenceCollection)
        .filter(EvidenceCollection.id == collection_uuid)
        .first()
    )
    if not collection:
        raise HTTPException(404, "Collection not found")
    if collection.is_system:
        raise HTTPException(403, "Cannot delete system collections")
    db.delete(collection)
    db.commit()
    return {"message": "Collection deleted successfully"}


async def add_to_collection_service(collection_id, evidence_id, db):
    try:
        collection_uuid = uuid.UUID(collection_id)
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid ID format")
    collection = (
        db.query(EvidenceCollection)
        .filter(EvidenceCollection.id == collection_uuid)
        .first()
    )
    if not collection:
        raise HTTPException(404, "Collection not found")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    existing = (
        db.query(EvidenceCollectionItem)
        .filter(
            and_(
                EvidenceCollectionItem.collection_id == collection_uuid,
                EvidenceCollectionItem.evidence_item_id == evidence_uuid,
            )
        )
        .first()
    )
    if existing:
        raise HTTPException(409, "Item already in collection")
    user = get_default_user(db)
    collection_item = EvidenceCollectionItem(
        collection_id=collection_uuid,
        evidence_item_id=evidence_uuid,
        added_method="manual",
        added_by=user.id,
    )
    db.add(collection_item)
    db.commit()
    return {"message": "Item added to collection"}


async def remove_from_collection_service(collection_id, evidence_id, db):
    try:
        collection_uuid = uuid.UUID(collection_id)
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid ID format")
    item = (
        db.query(EvidenceCollectionItem)
        .filter(
            and_(
                EvidenceCollectionItem.collection_id == collection_uuid,
                EvidenceCollectionItem.evidence_item_id == evidence_uuid,
            )
        )
        .first()
    )
    if not item:
        raise HTTPException(404, "Item not in collection")
    db.delete(item)
    db.commit()
    return {"message": "Item removed from collection"}


async def get_evidence_stats_service(db, case_id, project_id):
    cache_key = f"evidence:stats:{case_id or 'all'}:{project_id or 'all'}"
    cached = get_cached(cache_key)
    if cached:
        return cached

    # Build scope filters (case/project). For a Case linked to a Project, we include
    # both case_id and project_id items unless an explicit project_id is provided.
    scope_filters: list[Any] = []

    case_uuid: uuid.UUID | None = None
    project_uuid: uuid.UUID | None = None

    if case_id:
        try:
            case_uuid = uuid.UUID(str(case_id))
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail="Invalid case_id format"
            ) from exc

        linked_project_uuid = None
        if not project_id:
            try:
                case = db.query(Case).filter(Case.id == case_uuid).first()
                if case and getattr(case, "project_id", None):
                    linked_project_uuid = case.project_id
            except Exception:
                linked_project_uuid = None

        if linked_project_uuid:
            scope_filters.append(
                or_(
                    EvidenceItem.case_id == case_uuid,
                    EvidenceItem.project_id == linked_project_uuid,
                )
            )
        else:
            scope_filters.append(EvidenceItem.case_id == case_uuid)

    if project_id:
        try:
            project_uuid = uuid.UUID(str(project_id))
        except Exception as exc:
            raise HTTPException(
                status_code=400, detail="Invalid project_id format"
            ) from exc
        scope_filters.append(EvidenceItem.project_id == project_uuid)

    scope_query = db.query(EvidenceItem)
    if scope_filters:
        scope_query = scope_query.filter(*scope_filters)

    total = scope_query.count()
    image_exts = [
        "jpg",
        "jpeg",
        "png",
        "gif",
        "bmp",
        "webp",
        "tif",
        "tiff",
        "heic",
        "heif",
        "svg",
    ]
    image_types = ["image", "photo", "photograph"]
    image_count = scope_query.filter(
        or_(
            EvidenceItem.mime_type.ilike("image/%"),
            EvidenceItem.file_type.in_(image_exts),
            EvidenceItem.evidence_type.in_(image_types),
        )
    ).count()
    type_counts_raw = (
        scope_query.with_entities(
            EvidenceItem.evidence_type, func.count(EvidenceItem.id)
        )
        .group_by(EvidenceItem.evidence_type)
        .all()
    )
    status_counts_raw = (
        scope_query.with_entities(
            EvidenceItem.processing_status, func.count(EvidenceItem.id)
        )
        .group_by(EvidenceItem.processing_status)
        .all()
    )
    unassigned = (
        db.query(EvidenceItem)
        .filter(and_(EvidenceItem.case_id.is_(None), EvidenceItem.project_id.is_(None)))
        .count()
    )
    with_corr_query = db.query(
        func.count(func.distinct(EvidenceCorrespondenceLink.evidence_item_id))
    ).join(EvidenceItem, EvidenceItem.id == EvidenceCorrespondenceLink.evidence_item_id)
    if scope_filters:
        with_corr_query = with_corr_query.filter(*scope_filters)
    with_correspondence_raw = with_corr_query.scalar()
    with_correspondence = 0
    if isinstance(with_correspondence_raw, int):
        with_correspondence = with_correspondence_raw
    elif with_correspondence_raw is not None:
        with_correspondence = int(str(with_correspondence_raw))
    week_ago = datetime.now() - timedelta(days=7)
    recent = scope_query.filter(EvidenceItem.created_at >= week_ago).count()
    by_type = {}
    for row in type_counts_raw:
        type_val = row[0]
        count_val = row[1]
        if type_val is not None:
            count_int = 0
            if isinstance(count_val, int):
                count_int = count_val
            elif count_val is not None:
                count_int = int(str(count_val))
            by_type[str(type_val)] = count_int
    by_status = {}
    for row in status_counts_raw:
        status_val = row[0]
        status_count_val = row[1]
        if status_val is not None:
            status_count_int = 0
            if isinstance(status_count_val, int):
                status_count_int = status_count_val
            elif status_count_val is not None:
                status_count_int = int(str(status_count_val))
            by_status[str(status_val)] = status_count_int
    result = {
        "total": total,
        "unassigned": unassigned,
        "with_correspondence": with_correspondence,
        "recent_uploads": recent,
        "by_type": by_type,
        "by_status": by_status,
        "image_count": image_count,
    }
    set_cached(cache_key, result, ttl_seconds=60)
    return result


async def get_evidence_types_service():
    return {
        "evidence_types": [evidence_type.value for evidence_type in EvidenceType],
        "document_categories": [category.value for category in DocumentCategory],
        "link_types": [link_type.value for link_type in CorrespondenceLinkType],
        "relation_types": [
            relation_type.value for relation_type in EvidenceRelationType
        ],
    }


async def get_evidence_metadata_service(evidence_id, db, user, extract_fresh):
    from ..evidence_metadata import extract_evidence_metadata

    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    if item.extracted_metadata and not extract_fresh:
        return {
            "evidence_id": evidence_id,
            "metadata": item.extracted_metadata,
            "cached": True,
        }
    try:
        metadata = await extract_evidence_metadata(item.s3_key, item.s3_bucket)
        item.extracted_metadata = metadata
        item.metadata_extracted_at = datetime.now()
        db.commit()
        return {"evidence_id": evidence_id, "metadata": metadata, "cached": False}
    except Exception as e:
        logger.error(f"Error extracting metadata for {evidence_id}: {e}")
        return {
            "evidence_id": evidence_id,
            "metadata": {
                "extraction_status": "error",
                "extraction_error": str(e),
                "filename": item.filename,
                "file_size": item.file_size,
                "mime_type": item.mime_type,
            },
            "cached": False,
        }


async def get_evidence_preview_service(evidence_id, db, user):
    import mimetypes

    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    stored_mime_type = item.mime_type or "application/octet-stream"
    guessed_mime_type = mimetypes.guess_type(item.filename)[0]
    mime_type = stored_mime_type
    if stored_mime_type == "application/octet-stream" and guessed_mime_type:
        mime_type = guessed_mime_type
    preview_type = "unsupported"
    preview_url = None
    preview_content = None
    page_count = None
    dimensions = None
    try:
        preview_url = presign_get(
            item.s3_key,
            expires=3600,
            bucket=item.s3_bucket,
            response_content_type=mime_type,
        )
    except Exception as e:
        logger.warning(f"Could not generate presigned URL: {e}")
    if mime_type.startswith("image/"):
        preview_type = "image"
        if item.extracted_metadata:
            dimensions = {
                "width": item.extracted_metadata.get("width"),
                "height": item.extracted_metadata.get("height"),
            }
    elif mime_type == "application/pdf":
        preview_type = "pdf"
        if item.extracted_metadata:
            page_count = item.extracted_metadata.get("page_count")
    elif mime_type.startswith("text/") or item.filename.lower().endswith(
        (".txt", ".csv", ".json", ".xml", ".html", ".md", ".log")
    ):
        preview_type = "text"
        preview_content = "Text preview available via /text-content"
    elif mime_type.startswith("audio/"):
        preview_type = "audio"
    elif mime_type.startswith("video/"):
        preview_type = "video"
    elif mime_type in [
        "application/msword",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.ms-excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-powerpoint",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ]:
        preview_type = "office"
        if item.extracted_metadata:
            preview_content = item.extracted_metadata.get("text_preview")
            page_count = item.extracted_metadata.get(
                "page_count"
            ) or item.extracted_metadata.get("slide_count")
    elif mime_type in [
        "message/rfc822",
        "application/vnd.ms-outlook",
    ] or item.filename.lower().endswith((".eml", ".msg")):
        preview_type = "email"
        if item.extracted_metadata:
            preview_content = {
                "from": item.extracted_metadata.get("email_from"),
                "to": item.extracted_metadata.get("email_to"),
                "cc": item.extracted_metadata.get("email_cc"),
                "subject": item.extracted_metadata.get("email_subject"),
                "date": item.extracted_metadata.get("email_date"),
                "body_preview": item.extracted_metadata.get("text_preview"),
            }
    return {
        "evidence_id": evidence_id,
        "filename": item.filename,
        "mime_type": mime_type,
        "file_size": item.file_size,
        "preview_type": preview_type,
        "preview_url": preview_url,
        "preview_content": preview_content,
        "page_count": page_count,
        "dimensions": dimensions,
        "can_preview_inline": preview_type
        in ["image", "pdf", "text", "audio", "video"],
        "download_url": preview_url,
    }


async def get_evidence_office_render_service(
    evidence_id: str,
    db: Session,
    user: User,
    *,
    sheet: str | None = None,
    max_rows: int = 200,
    max_cols: int = 40,
    max_chars: int = 300_000,
) -> dict[str, Any]:
    """Render Office documents to lightweight HTML for in-app preview.

    Motivation: client-side previews (Mammoth/SheetJS) can fail due to CSP/CDN/race issues.
    This endpoint provides a server-side fallback using Python libs already in the image.
    """

    _ = user  # auth enforced by routes
    import html as _html
    import io
    import mimetypes
    import re

    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid ID format")

    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")

    filename = item.filename or ""
    lower_name = filename.lower()
    mime_type = (
        item.mime_type
        or mimetypes.guess_type(filename)[0]
        or "application/octet-stream"
    )

    # Hard caps to avoid huge responses.
    max_rows = max(1, min(int(max_rows or 200), 1000))
    max_cols = max(1, min(int(max_cols or 40), 200))
    max_chars = max(10_000, min(int(max_chars or 300_000), 2_000_000))

    data = get_object(item.s3_key, bucket=item.s3_bucket)
    if not data:
        raise HTTPException(404, "File content not found")

    def _wrap_html(body: str, *, title: str | None = None) -> str:
        safe_title = _html.escape(title or filename or "Document")
        return (
            "<!doctype html><html><head><meta charset='utf-8'/>"
            "<meta name='viewport' content='width=device-width, initial-scale=1'/>"
            f"<title>{safe_title}</title>"
            "<style>"
            "body{font-family:system-ui,-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;"
            "margin:0;padding:18px 20px;color:#111827;background:#fff;}"
            "h1,h2,h3{margin:0.8em 0 0.4em 0;}"
            "p{margin:0.6em 0;line-height:1.6;}"
            "table{border-collapse:collapse;width:100%;font-size:0.9rem;}"
            "th,td{border:1px solid #e5e7eb;padding:6px 10px;text-align:left;vertical-align:top;}"
            "th{background:#f3f4f6;position:sticky;top:0;z-index:1;}"
            "tr:nth-child(even){background:#fafafa;}"
            "pre{white-space:pre-wrap;word-break:break-word;}"
            "code{font-family:ui-monospace,SFMono-Regular,Menlo,Monaco,Consolas,monospace;}"
            "</style></head><body>"
            f"{body}"
            "</body></html>"
        )

    def _truncate(s: str) -> tuple[str, bool]:
        if len(s) <= max_chars:
            return s, False
        return s[:max_chars] + "\n<!-- truncated -->", True

    # ------------------------------------------------------------------
    # DOCX
    # ------------------------------------------------------------------
    if lower_name.endswith(".docx") or "wordprocessingml" in mime_type:
        try:
            from docx import Document  # type: ignore
            from docx.document import Document as _DocxDocument  # type: ignore
            from docx.oxml.table import CT_Tbl  # type: ignore
            from docx.oxml.text.paragraph import CT_P  # type: ignore
            from docx.table import Table  # type: ignore
            from docx.text.paragraph import Paragraph  # type: ignore
        except Exception as exc:
            raise HTTPException(
                500, f"DOCX preview support unavailable: {exc}"
            ) from exc

        def _iter_block_items(parent: _DocxDocument):
            parent_elm = parent.element.body
            for child in parent_elm.iterchildren():
                if isinstance(child, CT_P):
                    yield Paragraph(child, parent)
                elif isinstance(child, CT_Tbl):
                    yield Table(child, parent)

        def _render_paragraph(p: Paragraph) -> str:
            text = (p.text or "").strip()
            if not text and not p.runs:
                return ""
            # Basic run formatting.
            parts: list[str] = []
            for r in p.runs:
                t = r.text or ""
                if not t:
                    continue
                safe = _html.escape(t)
                if r.bold:
                    safe = f"<strong>{safe}</strong>"
                if r.italic:
                    safe = f"<em>{safe}</em>"
                if r.underline:
                    safe = f"<u>{safe}</u>"
                parts.append(safe)
            content = "".join(parts) if parts else _html.escape(p.text or "")
            content = content.strip()
            if not content:
                return ""

            style_name = None
            try:
                style_name = str(getattr(p.style, "name", "") or "")
            except Exception:
                style_name = ""

            if style_name.lower().startswith("heading"):
                m = re.search(r"(\d+)", style_name)
                level = 1
                if m:
                    try:
                        level = max(1, min(int(m.group(1)), 6))
                    except Exception:
                        level = 1
                return f"<h{level}>{content}</h{level}>"

            return f"<p>{content}</p>"

        def _render_table(t: Table) -> str:
            rows_html: list[str] = []
            for row in t.rows:
                cells_html: list[str] = []
                for cell in row.cells:
                    cells_html.append(
                        f"<td>{_html.escape(cell.text or '').strip()}</td>"
                    )
                rows_html.append("<tr>" + "".join(cells_html) + "</tr>")
            return "<table>" + "".join(rows_html) + "</table>"

        doc = Document(io.BytesIO(data))
        out_parts: list[str] = []
        for blk in _iter_block_items(doc):
            if hasattr(blk, "runs"):
                html_part = _render_paragraph(blk)  # type: ignore[arg-type]
            else:
                html_part = _render_table(blk)  # type: ignore[arg-type]
            if html_part:
                out_parts.append(html_part)
        html_body = _wrap_html("".join(out_parts), title=filename)
        html_body, truncated = _truncate(html_body)
        return {
            "evidence_id": evidence_id,
            "filename": filename,
            "mime_type": mime_type,
            "kind": "word",
            "sheet_names": [],
            "sheet": None,
            "html": html_body,
            "truncated": truncated,
        }

    # ------------------------------------------------------------------
    # XLSX / CSV
    # ------------------------------------------------------------------
    if lower_name.endswith((".xlsx", ".xlsm")) or "spreadsheetml" in mime_type:
        try:
            import openpyxl  # type: ignore
        except Exception as exc:
            raise HTTPException(
                500, f"XLSX preview support unavailable: {exc}"
            ) from exc

        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheet_names = list(wb.sheetnames or [])
        if not sheet_names:
            raise HTTPException(400, "Workbook contains no sheets")
        sheet_name = sheet if sheet in sheet_names else sheet_names[0]
        ws = wb[sheet_name]

        # Render a bounded grid.
        max_r = min(max_rows, int(getattr(ws, "max_row", max_rows) or max_rows))
        max_c = min(max_cols, int(getattr(ws, "max_column", max_cols) or max_cols))

        rows_html: list[str] = []
        for ridx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=max_r, max_col=max_c, values_only=True),
            start=1,
        ):
            cells = []
            tag = "th" if ridx == 1 else "td"
            for v in row:
                s = "" if v is None else str(v)
                cells.append(f"<{tag}>{_html.escape(s)}</{tag}>")
            rows_html.append("<tr>" + "".join(cells) + "</tr>")

        table = "<table>" + "".join(rows_html) + "</table>"
        html_body = _wrap_html(table, title=filename)
        html_body, truncated = _truncate(html_body)
        return {
            "evidence_id": evidence_id,
            "filename": filename,
            "mime_type": mime_type,
            "kind": "excel",
            "sheet_names": sheet_names,
            "sheet": sheet_name,
            "html": html_body,
            "truncated": truncated,
        }

    if lower_name.endswith(".csv") or mime_type in {"text/csv"}:
        # Lightweight CSV preview (bounded rows/cols).
        try:
            import csv
        except Exception:
            csv = None
        text = None
        try:
            text = data.decode("utf-8", errors="replace")
        except Exception:
            text = None
        if text is None or not csv:
            raise HTTPException(400, "CSV preview unavailable")

        reader = csv.reader(io.StringIO(text))
        rows = []
        for i, row in enumerate(reader):
            if i >= max_rows:
                break
            rows.append(row[:max_cols])
        if not rows:
            rows = [[]]

        rows_html = []
        for ridx, row in enumerate(rows, start=1):
            tag = "th" if ridx == 1 else "td"
            cells = [f"<{tag}>{_html.escape(str(v))}</{tag}>" for v in row]
            rows_html.append("<tr>" + "".join(cells) + "</tr>")
        table = "<table>" + "".join(rows_html) + "</table>"
        html_body = _wrap_html(table, title=filename)
        html_body, truncated = _truncate(html_body)
        return {
            "evidence_id": evidence_id,
            "filename": filename,
            "mime_type": mime_type,
            "kind": "excel",
            "sheet_names": ["CSV"],
            "sheet": "CSV",
            "html": html_body,
            "truncated": truncated,
        }

    raise HTTPException(
        400, "Office preview supported only for .docx, .xlsx, .xlsm, .csv"
    )


async def trigger_metadata_extraction_service(evidence_id, db, user):
    from ..evidence_metadata import extract_evidence_metadata

    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    try:
        metadata = await extract_evidence_metadata(item.s3_key, item.s3_bucket)
        item.extracted_metadata = metadata
        item.metadata_extracted_at = datetime.now()
        mime_raw = metadata.get("mime_type")
        if mime_raw is not None:
            mime = str(mime_raw)
            item.mime_type = mime
            if mime.startswith("image/"):
                item.evidence_type = "image"
            elif mime == "application/pdf":
                item.evidence_type = "pdf"
            elif "word" in mime or mime.endswith(".document"):
                item.evidence_type = "word_document"
            elif "excel" in mime or "spreadsheet" in mime:
                item.evidence_type = "spreadsheet"
        if not item.title and metadata.get("title"):
            item.title = metadata["title"]
        if metadata.get("author") and not item.author:
            item.author = metadata["author"]
        if metadata.get("page_count") and not item.page_count:
            item.page_count = metadata["page_count"]
        doc_date_val = _extract_metadata_date(metadata)
        email = None
        if item.source_email_id:
            email = (
                db.query(EmailMessage)
                .filter(EmailMessage.id == item.source_email_id)
                .first()
            )
        if doc_date_val is not None:
            item.document_date = doc_date_val
        elif (
            item.document_date is None
            and email
            and (email.date_sent or email.date_received)
        ):
            item.document_date = email.date_sent or email.date_received
        if email:
            domain_map, email_map = _load_stakeholder_maps(
                db, email.project_id, email.case_id
            )
            parties = _extract_company_parties(
                email.sender_email,
                email.recipients_to,
                email.recipients_cc,
                domain_map,
                email_map,
            )
            if parties:
                item.extracted_parties = _merge_parties(item.extracted_parties, parties)
        item.processing_status = "processed"
        item.processing_error = None
        db.commit()

        # Optional AWS-enhanced OCR/analysis for PDFs/images (Textract/Comprehend/etc.).
        # This upgrades evidence from "processed" metadata-only into a real "ready" state
        # with extracted_text when AWS services are configured.
        if (
            settings.USE_TEXTRACT
            and item.mime_type
            and (
                item.mime_type == "application/pdf"
                or item.mime_type.startswith("image/")
            )
        ):
            try:
                from ..enhanced_evidence_processor import enhanced_processor

                _ = await enhanced_processor.process_evidence_item(evidence_id, db)
            except Exception as aws_exc:
                logger.warning(
                    f"AWS enhanced evidence processing failed for {evidence_id}: {aws_exc}"
                )

        return {"evidence_id": evidence_id, "status": "completed", "metadata": metadata}
    except Exception as e:
        logger.error(f"Error extracting metadata for {evidence_id}: {e}")
        item.processing_status = "error"
        item.processing_error = str(e)
        db.commit()
        raise HTTPException(500, f"Metadata extraction failed: {str(e)}")


async def get_evidence_thumbnail_service(evidence_id, db, user, size):
    import mimetypes

    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")
    stored_mime_type = item.mime_type or "application/octet-stream"
    guessed_mime_type = mimetypes.guess_type(item.filename)[0]
    mime_type = stored_mime_type
    if stored_mime_type == "application/octet-stream" and guessed_mime_type:
        mime_type = guessed_mime_type
    size_map = {"small": 64, "medium": 200, "large": 400}
    thumb_size = size_map.get(size, 200)
    if mime_type.startswith("image/"):
        try:
            url = presign_get(
                item.s3_key,
                expires=3600,
                bucket=item.s3_bucket,
                response_content_type=mime_type,
            )
            return {
                "evidence_id": evidence_id,
                "thumbnail_type": "image",
                "thumbnail_url": url,
                "size": thumb_size,
                "original_dimensions": {
                    "width": item.extracted_metadata.get("width"),
                    "height": item.extracted_metadata.get("height"),
                },
            }
        except Exception as e:
            logger.warning(f"Could not generate thumbnail URL: {e}")
    icon_map = {
        "application/pdf": "pdf",
        "application/msword": "word",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "word",
        "application/vnd.ms-excel": "excel",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "excel",
        "application/vnd.ms-powerpoint": "powerpoint",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation": "powerpoint",
        "message/rfc822": "email",
        "application/vnd.ms-outlook": "email",
    }
    icon = "file"
    if mime_type in icon_map:
        icon = icon_map[mime_type]
    return {
        "evidence_id": evidence_id,
        "thumbnail_type": "placeholder",
        "icon": icon,
        "mime_type": mime_type,
        "filename": item.filename,
        "size": thumb_size,
    }


async def get_evidence_text_content_service(evidence_id, db, user, max_length):
    import httpx
    import os

    from .text_extract import extract_text_from_bytes, tika_url_candidates

    TIKA_URL = os.getenv("TIKA_URL", "http://tika:9998")
    try:
        evidence_uuid = uuid.UUID(evidence_id)
    except ValueError:
        raise HTTPException(400, "Invalid ID format")
    item = db.query(EvidenceItem).filter(EvidenceItem.id == evidence_uuid).first()
    if not item:
        raise HTTPException(404, "Evidence item not found")

    if item.extracted_text:
        return {
            "evidence_id": evidence_id,
            "text": item.extracted_text[:max_length],
            "total_length": len(item.extracted_text),
            "truncated": len(item.extracted_text) > max_length,
            "cached": True,
        }

    try:
        # Always use the item's bucket; evidence may be stored outside the default.
        content = get_object(item.s3_key, bucket=item.s3_bucket)
        if not content:
            raise HTTPException(404, "File content not found")
    except Exception as e:
        raise HTTPException(500, f"Could not retrieve file: {str(e)}")

    mime_type = item.mime_type or ""
    filename = item.filename or ""
    text = ""

    if mime_type.startswith("text/") or filename.endswith(
        (".txt", ".csv", ".json", ".xml", ".html", ".md")
    ):
        for encoding in ["utf-8", "latin-1", "cp1252"]:
            try:
                text = content.decode(encoding)
                break
            except Exception:
                continue
        if not text:
            text = content.decode("utf-8", errors="replace")
    else:
        # 1) Best-effort local extraction for Office docs (keeps UI working even
        #    if Tika is down or misconfigured).
        text = extract_text_from_bytes(
            content,
            filename=filename,
            mime_type=mime_type,
        )

        # 2) Fall back to Tika for everything else (and for legacy Office types).
        if not text.strip():
            tika_urls = tika_url_candidates(TIKA_URL)
            last_error: str | None = None
            for base in tika_urls:
                try:
                    async with httpx.AsyncClient(timeout=60.0) as client:
                        response = await client.put(
                            f"{base}/tika",
                            content=content,
                            headers={"Accept": "text/plain"},
                        )
                        if response.status_code == 200:
                            text = response.text
                            break
                        last_error = f"HTTP {response.status_code}"
                except httpx.TimeoutException:
                    last_error = "timeout"
                    continue
                except Exception as e:
                    last_error = str(e)
                    continue

            if not text.strip():
                # Don't persist a placeholder; return a helpful message so the
                # Evidence UI can still show *something*.
                tried = ", ".join(tika_urls)
                hint = (
                    "[Preview unavailable] Text extraction failed for this file type.\n"
                    f"Tika may be unreachable (tried: {tried}).\n"
                    "If you're running on Kubernetes, ensure TIKA_URL points to the in-cluster service (e.g. http://tika-service:9998).\n"
                    f"Last error: {last_error or 'unknown'}"
                )
                return {
                    "evidence_id": evidence_id,
                    "text": hint[:max_length],
                    "total_length": len(hint),
                    "truncated": len(hint) > max_length,
                    "cached": False,
                }

    item.extracted_text = text
    item.ocr_completed = True
    if item.processing_status != "error":
        item.processing_status = "ready"
        item.processing_error = None
        item.processed_at = datetime.utcnow()
    db.commit()
    return {
        "evidence_id": evidence_id,
        "text": text[:max_length],
        "total_length": len(text),
        "truncated": len(text) > max_length,
        "cached": False,
    }


async def sync_email_attachments_to_evidence_service(
    db, user, project_id, extract_metadata: bool = True
):
    import os
    from ..evidence_metadata import extract_evidence_metadata

    stakeholder_cache: dict[
        tuple[uuid.UUID | None, uuid.UUID | None], tuple[dict[str, str], dict[str, str]]
    ] = {}

    def _get_stakeholder_maps(project_uuid, case_uuid):
        key = (project_uuid, case_uuid)
        if key not in stakeholder_cache:
            stakeholder_cache[key] = _load_stakeholder_maps(db, project_uuid, case_uuid)
        return stakeholder_cache[key]

    existing_hashes = (
        db.query(EvidenceItem.file_hash)
        .filter(EvidenceItem.file_hash.isnot(None))
        .distinct()
        .all()
    )
    existing_hash_set = {h[0] for h in existing_hashes if h[0]}
    attachment_query = db.query(EmailAttachment).join(
        EmailMessage, EmailAttachment.email_message_id == EmailMessage.id
    )
    if project_id:
        try:
            project_uuid = uuid.UUID(project_id)
            attachment_query = attachment_query.filter(
                EmailMessage.project_id == project_uuid
            )
        except ValueError:
            raise HTTPException(400, "Invalid project_id format")
    attachments = attachment_query.filter(
        EmailAttachment.is_inline == False, EmailAttachment.s3_key.isnot(None)
    ).all()
    created_count = 0
    skipped_count = 0
    skipped_signature = 0  # Track signature images skipped
    error_count = 0
    metadata_extracted = 0
    metadata_errors = 0
    commit_interval = 25 if extract_metadata else 100

    # Patterns that indicate signature/inline images (skip these even if not marked as inline)
    signature_patterns = [
        "image00",
        "image0",
        "img00",
        "img0",  # Outlook auto-names: image001.png
        "signature",
        "logo",
        "banner",
        "footer",
        "linkedin",
        "twitter",
        "facebook",
        "instagram",  # Social media icons
        "email_signature",
        "emailsignature",
        "icon",
        "avatar",
        "cid:",
    ]
    image_exts = {"jpg", "jpeg", "png", "gif", "bmp", "webp", "tif", "tiff"}

    for att in attachments:
        try:
            if att.attachment_hash and att.attachment_hash in existing_hash_set:
                skipped_count += 1
                continue

            # Skip signature/inline images that weren't properly marked
            fname_lower = (att.filename or "").lower()
            raw_ext = (
                os.path.splitext(fname_lower)[1].lstrip(".") if fname_lower else ""
            )
            is_image = (
                att.content_type and att.content_type.startswith("image/")
            ) or raw_ext in image_exts
            is_signature_filename = any(
                pattern in fname_lower for pattern in signature_patterns
            )
            is_very_small = (att.file_size_bytes or 0) < 50000  # Under 50KB

            if is_image and (is_signature_filename or is_very_small):
                skipped_signature += 1
                logger.debug(
                    f"Skipping signature image: {att.filename} ({att.file_size_bytes} bytes)"
                )
                continue

            email = (
                db.query(EmailMessage)
                .filter(EmailMessage.id == att.email_message_id)
                .first()
            )
            if not email:
                error_count += 1
                continue
            email_date = email.date_sent or email.date_received
            domain_map, email_map = _get_stakeholder_maps(
                email.project_id, email.case_id
            )
            parties = _extract_company_parties(
                email.sender_email,
                email.recipients_to,
                email.recipients_cc,
                domain_map,
                email_map,
            )
            raw_ext = (
                os.path.splitext(att.filename or "")[1].lower().lstrip(".")
                if att.filename
                else ""
            )
            file_ext = (
                raw_ext
                if raw_ext and len(raw_ext) <= 10 and raw_ext.isalnum()
                else None
            )
            # Determine evidence_type based on actual content type, not just "email_attachment"
            # This ensures images go to Images & Media tab, PDFs are properly categorized, etc.
            mime = att.content_type or ""
            if mime.startswith("image/"):
                detected_evidence_type = "image"
            elif mime.startswith("video/"):
                detected_evidence_type = "video"
            elif mime.startswith("audio/"):
                detected_evidence_type = "audio"
            elif mime == "application/pdf":
                detected_evidence_type = "pdf"
            elif "word" in mime or mime.endswith(".document"):
                detected_evidence_type = "word_document"
            elif "excel" in mime or "spreadsheet" in mime:
                detected_evidence_type = "spreadsheet"
            elif "powerpoint" in mime or "presentation" in mime:
                detected_evidence_type = "presentation"
            else:
                # Fallback: detect from file extension
                image_exts = {
                    "jpg",
                    "jpeg",
                    "png",
                    "gif",
                    "bmp",
                    "webp",
                    "tif",
                    "tiff",
                    "heic",
                    "heif",
                    "svg",
                }
                video_exts = {"mp4", "mov", "m4v", "avi", "mkv", "webm", "wmv"}
                audio_exts = {"mp3", "wav", "m4a", "aac", "flac", "ogg", "opus"}
                if file_ext in image_exts:
                    detected_evidence_type = "image"
                elif file_ext in video_exts:
                    detected_evidence_type = "video"
                elif file_ext in audio_exts:
                    detected_evidence_type = "audio"
                elif file_ext == "pdf":
                    detected_evidence_type = "pdf"
                else:
                    detected_evidence_type = "document"  # Generic document type
            evidence_item = EvidenceItem(
                filename=att.filename or "unnamed_attachment",
                original_path=f"EmailAttachment:{att.id}",
                file_type=file_ext,
                mime_type=att.content_type,
                file_size=att.file_size_bytes,
                file_hash=att.attachment_hash,
                s3_bucket=att.s3_bucket or settings.S3_BUCKET,
                s3_key=att.s3_key,
                evidence_type=detected_evidence_type,
                source_type="pst_extraction",
                source_email_id=email.id,
                project_id=email.project_id,
                case_id=email.case_id,
                document_date=email_date,
                processing_status="pending",
                extracted_parties=parties,
                auto_tags=["email-attachment", "synced-from-attachments"],
            )
            db.add(evidence_item)
            if extract_metadata:
                try:
                    metadata = await extract_evidence_metadata(
                        evidence_item.s3_key, evidence_item.s3_bucket
                    )
                    evidence_item.extracted_metadata = metadata
                    evidence_item.metadata_extracted_at = datetime.now()
                    if metadata.get("title") and not evidence_item.title:
                        evidence_item.title = metadata["title"]
                    if metadata.get("author") and not evidence_item.author:
                        evidence_item.author = metadata["author"]
                    if metadata.get("page_count") and not evidence_item.page_count:
                        evidence_item.page_count = metadata["page_count"]
                    doc_date_val = _extract_metadata_date(metadata)
                    if doc_date_val:
                        if isinstance(doc_date_val, datetime):
                            evidence_item.document_date = (
                                doc_date_val.replace(tzinfo=None)
                                if doc_date_val.tzinfo
                                else doc_date_val
                            )
                        else:
                            evidence_item.document_date = datetime.combine(
                                doc_date_val, datetime.min.time()
                            )
                    evidence_item.processing_status = "processed"
                    metadata_extracted += 1
                except Exception as e:
                    metadata_errors += 1
                    evidence_item.processing_status = "error"
                    logger.error(
                        f"Metadata extraction failed for attachment {att.id}: {e}"
                    )
            if att.attachment_hash:
                existing_hash_set.add(att.attachment_hash)
            created_count += 1
            if created_count % commit_interval == 0:
                db.commit()
                logger.info(f"Synced {created_count} attachments so far...")
        except Exception as e:
            logger.error(f"Error syncing attachment {att.id}: {e}")
            error_count += 1
    db.commit()
    return {
        "status": "completed",
        "created": created_count,
        "skipped": skipped_count,
        "skipped_signature_images": skipped_signature,
        "errors": error_count,
        "metadata_extracted": metadata_extracted,
        "metadata_errors": metadata_errors,
        "total_processed": created_count
        + skipped_count
        + skipped_signature
        + error_count,
    }


async def get_sync_status_service(db, project_id):
    att_query = (
        db.query(func.count(EmailAttachment.id))
        .join(EmailMessage, EmailAttachment.email_message_id == EmailMessage.id)
        .filter(EmailAttachment.is_inline == False, EmailAttachment.s3_key.isnot(None))
    )
    if project_id:
        try:
            project_uuid = uuid.UUID(project_id)
            att_query = att_query.filter(EmailMessage.project_id == project_uuid)
        except ValueError:
            pass
    total_attachments = att_query.scalar() or 0
    ev_query = db.query(func.count(EvidenceItem.id)).filter(
        EvidenceItem.source_type == "pst_extraction"
    )
    if project_id:
        try:
            project_uuid = uuid.UUID(project_id)
            ev_query = ev_query.filter(EvidenceItem.project_id == project_uuid)
        except ValueError:
            pass
    total_evidence = ev_query.scalar() or 0
    missing_estimate = max(0, total_attachments - total_evidence)
    return {
        "total_attachments": total_attachments,
        "total_evidence_items": total_evidence,
        "missing_estimate": missing_estimate,
        "sync_needed": missing_estimate > 0,
    }


async def extract_all_metadata_service(db, user, limit, force):
    from ..evidence_metadata import extract_evidence_metadata

    # Bulk metadata extraction for EvidenceItems.
    # This endpoint is used to backfill items missing extracted_metadata and to
    # optionally refresh metadata when `force` is True.

    query = db.query(EvidenceItem)
    if not force:
        query = query.filter(
            or_(
                EvidenceItem.extracted_metadata.is_(None),
                EvidenceItem.metadata_extracted_at.is_(None),
            )
        )

    limit_int: int | None = None
    if limit is not None:
        try:
            limit_int = int(limit)
        except Exception:
            limit_int = None

    if limit_int is not None and limit_int > 0:
        query = query.order_by(EvidenceItem.created_at.desc()).limit(limit_int)

    items = query.all()

    processed = 0
    updated_dates = 0
    errors = 0

    email_cache: dict[uuid.UUID, EmailMessage | None] = {}

    def _get_email(email_id: uuid.UUID) -> EmailMessage | None:
        if email_id in email_cache:
            return email_cache[email_id]
        email_cache[email_id] = (
            db.query(EmailMessage).filter(EmailMessage.id == email_id).first()
        )
        return email_cache[email_id]

    for item in items:
        try:
            metadata = await extract_evidence_metadata(item.s3_key, item.s3_bucket)
            item.extracted_metadata = metadata
            item.metadata_extracted_at = datetime.now()
            doc_date_val = _extract_metadata_date(metadata)
            if doc_date_val is not None:
                if isinstance(doc_date_val, datetime):
                    item.document_date = (
                        doc_date_val.replace(tzinfo=None)
                        if doc_date_val.tzinfo
                        else doc_date_val
                    )
                else:
                    item.document_date = datetime.combine(
                        doc_date_val, datetime.min.time()
                    )
                updated_dates += 1
            email = None
            if item.source_email_id:
                email = _get_email(item.source_email_id)
            if (
                item.document_date is None
                and email
                and (email.date_sent or email.date_received)
            ):
                item.document_date = email.date_sent or email.date_received
                updated_dates += 1
            if metadata.get("author") and not item.author:
                item.author = metadata["author"]
            if metadata.get("page_count") and not item.page_count:
                item.page_count = metadata["page_count"]
            if metadata.get("title") and not item.title:
                item.title = metadata["title"]
            if email:
                domain_map, email_map = _load_stakeholder_maps(
                    db, email.project_id, email.case_id
                )
                parties = _extract_company_parties(
                    email.sender_email,
                    email.recipients_to,
                    email.recipients_cc,
                    domain_map,
                    email_map,
                )
                if parties:
                    item.extracted_parties = _merge_parties(
                        item.extracted_parties, parties
                    )
            item.processing_status = "processed"
            processed += 1
            if processed % 10 == 0:
                db.commit()
                logger.info(
                    f"Processed {processed} items, {updated_dates} dates updated..."
                )
        except Exception as e:
            logger.error(f"Error extracting metadata for {item.id}: {e}")
            errors += 1
    db.commit()
    return {
        "status": "completed",
        "processed": processed,
        "dates_updated": updated_dates,
        "errors": errors,
        "remaining": db.query(func.count(EvidenceItem.id))
        .filter(EvidenceItem.metadata_extracted_at.is_(None))
        .scalar()
        or 0,
    }
