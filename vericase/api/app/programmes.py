# pyright: reportMissingImports=false, reportMissingModuleSource=false
"""
Programme Management API
Handles upload and parsing of Asta Powerproject and PDF schedules
"""

from fastapi import (
    APIRouter,
    Depends,
    HTTPException,
    UploadFile,
    File,
    Form,
    BackgroundTasks,
)
from sqlalchemy.orm import Session
from typing import Any
import xml.etree.ElementTree as ET
from datetime import datetime, timezone
import io
import uuid
import openpyxl
import PyPDF2
import json
from dateutil import parser as date_parser

from .db import get_db
from .models import Programme, DelayEvent, Case, Document, User
from .security import current_user
from .ai_runtime import complete_chat
from .ai_settings import get_ai_api_key
from .delay_analysis import (
    DelayAnalysisOrchestrator,
    DelayAnalysisSession,
    _delay_sessions,
)

router = APIRouter()

# cSpell:ignore asta


def _extract_project_dates(project_elem: Any) -> tuple[str | None, str | None]:
    """Extract project start and finish dates from PROJECT element"""
    if project_elem is None:
        return None, None

    start_str = project_elem.get("START_DATE") or project_elem.get("start_date")
    finish_str = project_elem.get("FINISH_DATE") or project_elem.get("finish_date")

    project_start = parse_asta_date(start_str) if start_str else None
    project_finish = parse_asta_date(finish_str) if finish_str else None

    return project_start, project_finish


def _parse_task_element(task: Any) -> dict[str, Any]:
    """Parse a single TASK element into activity dict"""
    return {
        "id": task.get("ID") or task.get("id"),
        "name": task.get("NAME") or task.get("name") or task.text,
        "start_date": parse_asta_date(task.get("START_DATE") or task.get("start_date")),
        "finish_date": parse_asta_date(
            task.get("FINISH_DATE") or task.get("finish_date")
        ),
        "duration": task.get("DURATION") or task.get("duration") or "0",
        "percent_complete": task.get("PERCENT_COMPLETE")
        or task.get("percent_complete")
        or "0",
        "is_critical": task.get("CRITICAL") == "true" or task.get("critical") == "true",
        "is_milestone": task.get("MILESTONE") == "true"
        or task.get("milestone") == "true",
    }


def _calculate_project_dates(
    activities: list[dict[str, Any]],
) -> tuple[str | None, str | None]:
    """Calculate project start/finish from activities if not provided"""
    if not activities:
        return None, None

    start_dates = [a["start_date"] for a in activities if a["start_date"]]
    finish_dates = [a["finish_date"] for a in activities if a["finish_date"]]

    project_start = min(start_dates) if start_dates else None
    project_finish = max(finish_dates) if finish_dates else None

    return project_start, project_finish


def parse_asta_xml(xml_content: bytes) -> dict[str, Any]:
    """
    Parse Asta Powerproject XML export to extract activities, dates, and critical path

    Asta XML structure typically includes:
    - PROJECT element with metadata
    - TASK elements with IDs, names, start/finish dates
    - LINKS for dependencies
    - EXCEPTIONS_CALENDAR for working days
    """
    try:
        root = ET.fromstring(xml_content)

        activities = []
        critical_path = []
        milestones = []

        # Extract project dates
        project_start, project_finish = _extract_project_dates(root.find(".//PROJECT"))

        # Parse tasks/activities
        for task in root.findall(".//TASK"):
            activity = _parse_task_element(task)
            activities.append(activity)

            if activity["is_critical"]:
                critical_path.append(activity["id"])

            if activity["is_milestone"]:
                milestones.append(
                    {
                        "id": activity["id"],
                        "name": activity["name"],
                        "date": activity["start_date"],
                    }
                )

        # Calculate dates from activities if not found
        if not project_start or not project_finish:
            calc_start, calc_finish = _calculate_project_dates(activities)
            project_start = project_start or calc_start
            project_finish = project_finish or calc_finish

        return {
            "activities": activities,
            "critical_path": critical_path,
            "milestones": milestones,
            "project_start": project_start,
            "project_finish": project_finish,
            "total_activities": len(activities),
        }

    except ET.ParseError as e:
        raise ValueError(f"Invalid XML format: {str(e)}")


def parse_asta_date(date_str: str | None) -> str | None:
    """
    Parse Asta date string to ISO format
    Asta typically uses: YYYY-MM-DD or DD/MM/YYYY
    """
    if not date_str:
        return None

    # Try ISO format first
    try:
        dt = datetime.fromisoformat(date_str)
        return dt.isoformat()
    except (ValueError, TypeError):
        pass

    # Try DD/MM/YYYY
    try:
        dt = datetime.strptime(date_str, "%d/%m/%Y")
        return dt.isoformat()
    except (ValueError, TypeError):
        pass

    # Try other common formats
    for fmt in ["%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]:
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.isoformat()
        except (ValueError, TypeError):
            continue

    return date_str  # Return as-is if can't parse


async def parse_pdf_schedule(pdf_content: bytes, db: Session) -> dict[str, Any]:
    """
    Extract schedule data from PDF using LLM
    """
    try:
        # Extract text from PDF
        pdf_file = io.BytesIO(pdf_content)
        reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"

        if not text.strip():
            return {
                "activities": [],
                "critical_path": [],
                "milestones": [],
                "project_start": None,
                "project_finish": None,
                "total_activities": 0,
                "note": "Could not extract text from PDF (scanned?)",
            }

        # Use LLM to parse activities
        prompt = f"""Extract construction schedule activities from this text.
        Return a JSON object with this structure:
        {{
            "activities": [
                {{
                    "id": "activity_id",
                    "name": "activity_name",
                    "start_date": "YYYY-MM-DD",
                    "finish_date": "YYYY-MM-DD",
                    "is_milestone": boolean
                }}
            ],
            "project_start": "YYYY-MM-DD",
            "project_finish": "YYYY-MM-DD"
        }}

        Text:
        {text[:10000]}  # Limit context
        """

        api_key = get_ai_api_key("openai", db) or get_ai_api_key("anthropic", db)
        if not api_key:
            return {
                "activities": [],
                "critical_path": [],
                "milestones": [],
                "project_start": None,
                "project_finish": None,
                "total_activities": 0,
                "note": "AI service not configured for PDF parsing",
            }

        provider = "openai" if get_ai_api_key("openai", db) else "anthropic"
        model = "gpt-4o" if provider == "openai" else "claude-3-5-sonnet-20240620"

        response = await complete_chat(
            provider=provider,
            model_id=model,
            prompt=prompt,
            system_prompt="You are a data extraction assistant. Extract schedule data accurately.",
            db=db,
            max_tokens=4000,
        )

        # Parse JSON response
        json_str = response
        if "```json" in response:
            json_str = response.split("```json")[1].split("```")[0]
        elif "```" in response:
            json_str = response.split("```")[1].split("```")[0]

        data = json.loads(json_str.strip())

        return {
            "activities": data.get("activities", []),
            "critical_path": [],
            "milestones": [
                a for a in data.get("activities", []) if a.get("is_milestone")
            ],
            "project_start": data.get("project_start"),
            "project_finish": data.get("project_finish"),
            "total_activities": len(data.get("activities", [])),
            "note": "Parsed via AI from PDF",
        }

    except Exception as e:
        return {
            "activities": [],
            "critical_path": [],
            "milestones": [],
            "project_start": None,
            "project_finish": None,
            "total_activities": 0,
            "note": f"PDF parsing failed: {str(e)}",
        }


def parse_excel_programme(file_content: bytes) -> dict[str, Any]:
    """
    Parse Excel schedule.
    Expected columns: Activity ID, Activity Name, Start Date, Finish Date
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active

        if ws is None:
            raise HTTPException(
                status_code=400,
                detail="Invalid Excel file: no active worksheet found",
            )

        activities = []
        milestones = []

        # Identify headers
        headers = {}
        header_row = 1
        for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            for idx, cell in enumerate(row):
                if isinstance(cell, str):
                    lower_cell = cell.lower()
                    if "id" in lower_cell and "activity" in lower_cell:
                        headers["id"] = idx
                    elif (
                        "name" in lower_cell
                        or "task" in lower_cell
                        or "activity" in lower_cell
                    ):
                        headers["name"] = idx
                    elif "start" in lower_cell:
                        headers["start"] = idx
                    elif "finish" in lower_cell or "end" in lower_cell:
                        headers["finish"] = idx
            if len(headers) >= 3:  # Found enough headers
                break
            header_row += 1

        if not headers:
            # Fallback to standard columns A, B, C, D
            headers = {"id": 0, "name": 1, "start": 2, "finish": 3}

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row[headers.get("name", 1)]:  # Skip empty rows
                continue

            start_val = row[headers.get("start", 2)]
            finish_val = row[headers.get("finish", 3)]

            start_date = None
            finish_date = None

            if isinstance(start_val, datetime):
                start_date = start_val.isoformat()
            elif start_val:
                try:
                    start_date = date_parser.parse(str(start_val)).isoformat()
                except Exception:
                    pass

            if isinstance(finish_val, datetime):
                finish_date = finish_val.isoformat()
            elif finish_val:
                try:
                    finish_date = date_parser.parse(str(finish_val)).isoformat()
                except Exception:
                    pass

            activity = {
                "id": (
                    str(row[headers.get("id", 0)])
                    if row[headers.get("id", 0)]
                    else str(uuid.uuid4())[:8]
                ),
                "name": str(row[headers.get("name", 1)]),
                "start_date": start_date,
                "finish_date": finish_date,
                "duration": "0",  # Calculate if needed
                "percent_complete": "0",
                "is_critical": False,
                "is_milestone": False,
            }

            if start_date and finish_date and start_date == finish_date:
                activity["is_milestone"] = True
                milestones.append(
                    {"id": activity["id"], "name": activity["name"], "date": start_date}
                )

            activities.append(activity)

        # Calculate project dates
        start_dates = [a["start_date"] for a in activities if a["start_date"]]
        finish_dates = [a["finish_date"] for a in activities if a["finish_date"]]

        project_start = min(start_dates) if start_dates else None
        project_finish = max(finish_dates) if finish_dates else None

        return {
            "activities": activities,
            "critical_path": [],
            "milestones": milestones,
            "project_start": project_start,
            "project_finish": project_finish,
            "total_activities": len(activities),
        }
    except Exception as e:
        raise ValueError(f"Invalid Excel format: {str(e)}")


@router.post("/api/programmes/upload")
async def upload_programme(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    case_id: uuid.UUID | None = Form(None),
    project_id: uuid.UUID | None = Form(None),
    programme_type: str = Form(...),  # as_planned, as_built, interim
    programme_date: str = Form(...),
    version_number: str | None = Form(None),
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """
    Upload and parse Asta Powerproject or PDF programme

    Supported formats:
    - .xml (Asta XML export - recommended)
    - .pp (Asta Powerproject - requires XML export first)
    - .pdf (PDF schedule - limited parsing)
    """
    from .models import Project

    # Ensure at least one context is provided
    if not case_id and not project_id:
        raise HTTPException(
            status_code=400, detail="Either case_id or project_id is required"
        )

    # Verify context exists
    _context_id = case_id or project_id  # noqa: F841
    if case_id:
        case = db.query(Case).filter(Case.id == case_id).first()
        if not case:
            raise HTTPException(status_code=404, detail="Case not found")
        # If case has a project, use that project_id
        if not project_id and hasattr(case, "project_id") and case.project_id:
            project_id = case.project_id
    elif project_id:
        project = db.query(Project).filter(Project.id == project_id).first()
        if not project:
            raise HTTPException(status_code=404, detail="Project not found")

    # Read file content
    content = await file.read()

    # Determine file type and parse
    if not file.filename:
        raise HTTPException(
            status_code=400, detail="Uploaded file must have a filename."
        )
    filename_lower = file.filename.lower()

    if filename_lower.endswith(".xml"):
        parsed_data = parse_asta_xml(content)
    elif filename_lower.endswith(".pp"):
        raise HTTPException(
            status_code=400,
            detail="Asta .pp files must be exported to XML format first. In Asta Powerproject: File > Export > XML",
        )
    elif filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
        parsed_data = parse_excel_programme(content)
    elif filename_lower.endswith(".pdf"):
        parsed_data = await parse_pdf_schedule(content, db)
    else:
        raise HTTPException(
            status_code=400,
            detail="Unsupported file format. Please upload .xml (Asta), .xlsx (Excel), or .pdf",
        )

    try:
        # Determine storage key prefix based on context
        storage_prefix = f"programmes/{case_id or project_id}"

        # Create document record
        document = Document(
            filename=file.filename,
            size=len(content),
            content_type=file.content_type or "application/octet-stream",
            bucket="vericase-documents",
            s3_key=f"{storage_prefix}/{file.filename}",
            owner_user_id=user.id,
        )
        db.add(document)
        db.flush()

        # Create programme record
        programme = Programme(
            case_id=case_id,
            project_id=project_id,
            programme_name=file.filename,
            programme_type=programme_type,
            programme_date=(
                datetime.fromisoformat(programme_date)
                if programme_date
                else datetime.now(timezone.utc)
            ),
            version_number=version_number,
            activities=parsed_data["activities"],
            critical_path=parsed_data["critical_path"],
            milestones=parsed_data["milestones"],
            project_start=(
                datetime.fromisoformat(parsed_data["project_start"])
                if parsed_data.get("project_start")
                else None
            ),
            project_finish=(
                datetime.fromisoformat(parsed_data["project_finish"])
                if parsed_data.get("project_finish")
                else None
            ),
            notes=f"Uploaded by {user.email}. {parsed_data['total_activities']} activities parsed.",
            filename=file.filename,
            s3_bucket="vericase-programmes",
            s3_key=f"{storage_prefix}/{file.filename}",
            file_format=file.filename.split(".")[-1].upper(),
            uploaded_by=user.id,
        )
        db.add(programme)
        db.commit()
    except (ValueError, TypeError, AttributeError) as e:
        import logging

        logging.error(f"Error saving programme data: {e}", exc_info=True)
        db.rollback()
        raise HTTPException(status_code=500, detail="Failed to save programme data")
    db.refresh(programme)

    # Auto-trigger email-to-activity linking in background
    # This populates as_planned_activity, as_built_activity, delay_days in EmailMessage
    if parsed_data["total_activities"] > 0:
        from .tasks import link_emails_to_programme_activities_task

        if project_id:
            background_tasks.add_task(
                link_emails_to_programme_activities_task.delay,
                project_id=str(project_id),
                overwrite_existing=False,
            )
        elif case_id:
            background_tasks.add_task(
                link_emails_to_programme_activities_task.delay,
                case_id=str(case_id),
                overwrite_existing=False,
            )

    return {
        "programme_id": programme.id,
        "document_id": document.id,
        "activities_parsed": parsed_data["total_activities"],
        "critical_activities": len(parsed_data["critical_path"]),
        "milestones": len(parsed_data["milestones"]),
        "project_start": parsed_data.get("project_start"),
        "project_finish": parsed_data.get("project_finish"),
        "linking_triggered": parsed_data["total_activities"] > 0,
    }


@router.post("/api/projects/{project_id}/programmes/link-emails")
async def trigger_project_email_linking(
    project_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    overwrite: bool = False,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """
    Manually trigger email-to-programme activity linking for a project.

    This links each email to its corresponding programme activity based on date,
    populating as_planned_activity, as_built_activity, and delay_days fields.
    """
    from .models import Project
    from .tasks import link_emails_to_programme_activities_task

    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    # Check if there are programmes to link against
    programme_count = (
        db.query(Programme).filter(Programme.project_id == project_id).count()
    )
    if programme_count == 0:
        raise HTTPException(
            status_code=400,
            detail="No programmes found for this project. Upload a programme first.",
        )

    # Trigger background task
    background_tasks.add_task(
        link_emails_to_programme_activities_task.delay,
        project_id=str(project_id),
        overwrite_existing=overwrite,
    )

    return {
        "status": "triggered",
        "project_id": str(project_id),
        "programmes_available": programme_count,
        "overwrite_existing": overwrite,
        "message": "Email linking task queued. Check correspondence for updated activity fields.",
    }


@router.post("/api/cases/{case_id}/programmes/link-emails")
async def trigger_case_email_linking(
    case_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    overwrite: bool = False,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """
    Manually trigger email-to-programme activity linking for a case.

    This links each email to its corresponding programme activity based on date,
    populating as_planned_activity, as_built_activity, and delay_days fields.
    """
    from .tasks import link_emails_to_programme_activities_task

    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    # Check if there are programmes to link against
    programme_count = db.query(Programme).filter(Programme.case_id == case_id).count()
    if programme_count == 0:
        raise HTTPException(
            status_code=400,
            detail="No programmes found for this case. Upload a programme first.",
        )

    # Trigger background task
    background_tasks.add_task(
        link_emails_to_programme_activities_task.delay,
        case_id=str(case_id),
        overwrite_existing=overwrite,
    )

    return {
        "status": "triggered",
        "case_id": str(case_id),
        "programmes_available": programme_count,
        "overwrite_existing": overwrite,
        "message": "Email linking task queued. Check correspondence for updated activity fields.",
    }


@router.get("/api/programmes/{programme_id}")
async def get_programme(
    programme_id: uuid.UUID,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """Get programme details"""
    try:
        programme = db.query(Programme).filter(Programme.id == programme_id).first()
        if not programme:
            raise HTTPException(status_code=404, detail="Programme not found")
    except HTTPException:
        raise
    except Exception as e:
        import logging

        logging.error(f"Error fetching programme: {e}")
        raise HTTPException(status_code=500, detail="Failed to fetch programme")

    return {
        "id": programme.id,
        "case_id": programme.case_id,
        "programme_type": programme.programme_type,
        "programme_date": (
            programme.programme_date.isoformat() if programme.programme_date else None
        ),
        "version_number": programme.version_number,
        "project_start": (
            programme.project_start.isoformat() if programme.project_start else None
        ),
        "project_finish": (
            programme.project_finish.isoformat() if programme.project_finish else None
        ),
        "activities": programme.activities,
        "critical_path": programme.critical_path,
        "milestones": programme.milestones,
        "total_activities": len(programme.activities) if programme.activities else 0,
        "notes": programme.notes,
    }


@router.get("/api/cases/{case_id}/programmes")
async def list_case_programmes(
    case_id: uuid.UUID,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """List all programmes for a case"""
    programmes = (
        db.query(Programme)
        .filter(Programme.case_id == case_id)
        .order_by(Programme.programme_date.desc())
        .all()
    )

    return [
        {
            "id": p.id,
            "programme_name": p.programme_name,
            "programme_type": p.programme_type,
            "programme_date": (
                p.programme_date.isoformat() if p.programme_date else None
            ),
            "version_number": p.version_number,
            "total_activities": len(p.activities) if p.activities else 0,
            "project_start": p.project_start.isoformat() if p.project_start else None,
            "project_finish": (
                p.project_finish.isoformat() if p.project_finish else None
            ),
        }
        for p in programmes
    ]


@router.get("/api/projects/{project_id}/programmes")
async def list_project_programmes(
    project_id: uuid.UUID,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """List all programmes for a project (including those from linked cases)"""
    from sqlalchemy import or_

    programmes = (
        db.query(Programme)
        .filter(
            or_(
                Programme.project_id == project_id,
                # Also include programmes from cases linked to this project
                Programme.case_id.in_(
                    db.query(Case.id).filter(Case.project_id == project_id)
                ),
            )
        )
        .order_by(Programme.programme_date.desc())
        .all()
    )

    return [
        {
            "id": p.id,
            "programme_name": p.programme_name,
            "programme_type": p.programme_type,
            "programme_date": (
                p.programme_date.isoformat() if p.programme_date else None
            ),
            "version_number": p.version_number,
            "total_activities": len(p.activities) if p.activities else 0,
            "project_start": p.project_start.isoformat() if p.project_start else None,
            "project_finish": (
                p.project_finish.isoformat() if p.project_finish else None
            ),
            "case_id": str(p.case_id) if p.case_id else None,
            "project_id": str(p.project_id) if p.project_id else None,
        }
        for p in programmes
    ]


@router.post("/api/programmes/compare")
async def compare_programmes(
    as_planned_id: uuid.UUID,
    as_built_id: uuid.UUID,
    background_tasks: BackgroundTasks,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """
    Compare as-planned vs as-built programmes to identify delays

    Returns:
    - Delays on critical path
    - Total float consumed
    - Activities with slippage
    """
    as_planned = db.query(Programme).filter(Programme.id == as_planned_id).first()
    as_built = db.query(Programme).filter(Programme.id == as_built_id).first()

    if not as_planned or not as_built:
        raise HTTPException(status_code=404, detail="Programme not found")

    if as_planned.case_id != as_built.case_id:
        raise HTTPException(
            status_code=400, detail="Programmes must be from the same case"
        )

    delays = []
    critical_delays = []

    # Match activities by ID
    planned_activities = {a["id"]: a for a in (as_planned.activities or [])}
    built_activities = {a["id"]: a for a in (as_built.activities or [])}

    for activity_id, planned in planned_activities.items():
        built = built_activities.get(activity_id)
        if not built:
            continue

        # Calculate delay
        planned_start = planned.get("start_date")
        planned_finish = planned.get("finish_date")
        built_start = built.get("start_date")
        built_finish = built.get("finish_date")

        if planned_finish and built_finish:
            try:
                planned_dt = datetime.fromisoformat(planned_finish)
                built_dt = datetime.fromisoformat(built_finish)
                delay_days = (built_dt - planned_dt).days

                if delay_days != 0:
                    is_critical = activity_id in (as_planned.critical_path or [])

                    delay_info = {
                        "activity_id": activity_id,
                        "activity_name": planned["name"],
                        "planned_finish": planned_finish,
                        "actual_finish": built_finish,
                        "delay_days": delay_days,
                        "is_critical": is_critical,
                    }

                    delays.append(delay_info)

                    if is_critical:
                        critical_delays.append(delay_info)

                        # Create delay event record
                        delay_event = DelayEvent(
                            case_id=as_planned.case_id,
                            as_planned_programme_id=as_planned.id,
                            as_built_programme_id=as_built.id,
                            activity_id=activity_id,
                            activity_name=planned["name"],
                            planned_start=(
                                datetime.fromisoformat(planned_start)
                                if planned_start
                                else None
                            ),
                            planned_finish=datetime.fromisoformat(planned_finish),
                            actual_start=(
                                datetime.fromisoformat(built_start)
                                if built_start
                                else None
                            ),
                            actual_finish=datetime.fromisoformat(built_finish),
                            delay_days=delay_days,
                            is_on_critical_path=True,
                            delay_type="critical",
                            description="Auto-detected from programme comparison",
                        )
                        db.add(delay_event)
            except (ValueError, TypeError):
                continue

    db.commit()

    # Auto-trigger delay analysis for significant critical delays
    significant_delays = [d for d in critical_delays if d["delay_days"] > 5]
    if significant_delays:
        try:
            # Create session
            session_id = str(uuid.uuid4())
            session = DelayAnalysisSession(
                id=session_id,
                user_id=str(user.id),
                case_id=str(as_planned.case_id),
                status="pending",
            )
            _delay_sessions[session_id] = session

            # Prepare events for analysis
            events_to_analyze = []
            for d in significant_delays[:5]:  # Limit to top 5
                events_to_analyze.append(
                    {
                        "description": f"Delay to {d['activity_name']} (ID: {d['activity_id']})",
                        "event_date": d.get("actual_finish"),
                        "planned_date": d.get("planned_finish"),
                        "actual_date": d.get("actual_finish"),
                        "delay_days": d["delay_days"],
                        "evidence_ids": [],
                    }
                )

            # Run in background
            async def run_analysis():
                orchestrator = DelayAnalysisOrchestrator(db, session)
                await orchestrator.analyze_delays(events_to_analyze)

            background_tasks.add_task(run_analysis)

        except Exception as e:
            import logging

            logging.error(f"Failed to auto-trigger delay analysis: {e}")

    return {
        "case_id": as_planned.case_id,
        "as_planned_programme": as_planned_id,
        "as_built_programme": as_built_id,
        "total_delays": len(delays),
        "critical_delays_count": len(critical_delays),
        "delays": (
            sorted(delays, key=lambda x: abs(x["delay_days"]), reverse=True)[:50]
            if delays
            else []
        ),  # Top 50
        "critical_delays": critical_delays,
        "summary": {
            "longest_delay": max([d["delay_days"] for d in delays]) if delays else 0,
            "total_delay_days": sum([d["delay_days"] for d in critical_delays]),
            "activities_delayed": len([d for d in delays if d["delay_days"] > 0]),
        },
    }


@router.get("/api/cases/{case_id}/delays")
async def list_delays(
    case_id: uuid.UUID,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """List all delay events for a case"""
    delays = (
        db.query(DelayEvent)
        .filter(DelayEvent.case_id == case_id)
        .order_by(DelayEvent.delay_days.desc())
        .all()
    )

    return [
        {
            "id": d.id,
            "activity_name": d.activity_name,
            "delay_days": d.delay_days,
            "delay_type": d.delay_type,
            "delay_cause": d.delay_cause,
            "is_on_critical_path": d.is_on_critical_path,
            "planned_finish": (
                d.planned_finish.isoformat() if d.planned_finish else None
            ),
            "actual_finish": d.actual_finish.isoformat() if d.actual_finish else None,
            "linked_correspondence": d.linked_correspondence_ids,
            "notes": d.notes,
        }
        for d in delays
    ]


from pydantic import BaseModel
from typing import Optional


class DelayEventCreate(BaseModel):
    """Create a new delay event"""

    activity_name: str
    delay_days: int = 0
    delay_type: str = "critical"  # critical, non_critical
    delay_cause: str = "neutral"  # employer, contractor, neutral, concurrent
    is_on_critical_path: bool = True
    planned_finish: Optional[str] = None
    actual_finish: Optional[str] = None
    notes: Optional[str] = None
    description: Optional[str] = None


@router.post("/api/cases/{case_id}/delays")
async def create_delay(
    case_id: uuid.UUID,
    delay_data: DelayEventCreate,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """Create a new delay event for a case"""
    # Verify case exists
    case = db.query(Case).filter(Case.id == case_id).first()
    if not case:
        raise HTTPException(status_code=404, detail="Case not found")

    # Parse dates
    planned = None
    actual = None
    if delay_data.planned_finish:
        try:
            planned = date_parser.parse(delay_data.planned_finish).replace(tzinfo=None)
        except Exception:
            pass
    if delay_data.actual_finish:
        try:
            actual = date_parser.parse(delay_data.actual_finish).replace(tzinfo=None)
        except Exception:
            pass

    delay = DelayEvent(
        case_id=case_id,
        activity_name=delay_data.activity_name,
        delay_days=delay_data.delay_days,
        delay_type=delay_data.delay_type,
        delay_cause=delay_data.delay_cause,
        is_on_critical_path=delay_data.is_on_critical_path,
        planned_finish=planned,
        actual_finish=actual,
        notes=delay_data.notes,
        description=delay_data.description,
    )

    db.add(delay)
    db.commit()
    db.refresh(delay)

    return {
        "id": delay.id,
        "activity_name": delay.activity_name,
        "delay_days": delay.delay_days,
        "delay_type": delay.delay_type,
        "delay_cause": delay.delay_cause,
        "is_on_critical_path": delay.is_on_critical_path,
        "planned_finish": (
            delay.planned_finish.isoformat() if delay.planned_finish else None
        ),
        "actual_finish": (
            delay.actual_finish.isoformat() if delay.actual_finish else None
        ),
        "notes": delay.notes,
        "message": "Delay event created successfully",
    }


@router.patch("/api/delays/{delay_id}/link-correspondence")
async def link_correspondence_to_delay(
    delay_id: int,
    evidence_ids: list[int],
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """Link correspondence (evidence) to a delay event"""
    delay = db.query(DelayEvent).filter(DelayEvent.id == delay_id).first()
    if not delay:
        raise HTTPException(status_code=404, detail="Delay event not found")

    # Update linked correspondence
    current_links = delay.linked_correspondence_ids or []
    current_links_str = [str(x) for x in current_links]
    incoming_str = [str(x) for x in evidence_ids]
    new_links = list(dict.fromkeys(current_links_str + incoming_str))

    delay.linked_correspondence_ids = new_links
    db.commit()

    return {
        "delay_id": delay_id,
        "linked_correspondence": new_links,
        "total_linked": len(new_links),
    }


@router.get("/api/programmes/{programme_id}/active-activities")
async def get_active_activities(
    programme_id: uuid.UUID,
    date: str,
    db: Session = Depends(get_db),
    user: "User" = Depends(current_user),
):
    """
    Get activities that were active on a specific date.
    Useful for auto-linking correspondence to activities.
    """
    programme = db.query(Programme).filter(Programme.id == programme_id).first()
    if not programme:
        raise HTTPException(status_code=404, detail="Programme not found")

    try:
        target_date = datetime.fromisoformat(date).replace(tzinfo=None)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format")

    active_activities = []

    if not programme.activities:
        return []

    for activity in programme.activities:
        start_str = activity.get("start_date")
        finish_str = activity.get("finish_date")

        if start_str and finish_str:
            try:
                start = datetime.fromisoformat(start_str).replace(tzinfo=None)
                finish = datetime.fromisoformat(finish_str).replace(tzinfo=None)

                # Check if target date is within range (inclusive)
                if start <= target_date <= finish:
                    active_activities.append(activity)
            except (ValueError, TypeError):
                continue

    return active_activities
